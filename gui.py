import sys
import customtkinter as ctk
from tkinter import filedialog, messagebox
import tkinter as tk
import threading
from datetime import datetime
import os
import traceback

from steps.merge_ekin_apel import merge_ekin_apel
from steps.excel_csv import excel_sheet_disiplin_ke_csv
from steps.download import download_rekap, stop_event, stop_download
from steps.input_spesimen import input_spesimen
from steps.analisis import analisis_kehadiran
from steps.merge2 import process_mail_merge
from steps.merge_pdf_jabatan import merge_pdf_by_jabatan
from steps.split_pdf_jabatan import split_pdf_jabatan
from steps.rekap_apel import rekap_kehadiran_apel
from config import TEMPLATE_WORD
from services.firebase_access import FirebaseRealtimeUserAccessGateway, ValidateUserAccess

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    from tkinterdnd2 import DND_FILES
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False

if getattr(sys, 'frozen', False):
    os.environ['WDM_LOCAL'] = '1'

FIREBASE_URL = "https://simak-tpp-default-rtdb.asia-southeast1.firebasedatabase.app"
APP_USERNAME = "dinaskehutanan"

# ══════════════════════════════════════════════════════════
# Color Palette — Dinas Kehutanan Provinsi Lampung
# ══════════════════════════════════════════════════════════
COLORS = {
    "sidebar_bg":       "#1B4332",
    "sidebar_hover":    "#2D6A4F",
    "sidebar_text":     "#FFFFFF",
    "sidebar_subtext":  "#95D5B2",

    "accent_green":     "#2D6A4F",
    "accent_green_hover": "#40916C",
    "accent_light":     "#B7E4C7",

    "content_bg":       "#F0F2F5",
    "card_bg":          "#FFFFFF",
    "card_border":      "#E0E0E0",

    "text_primary":     "#1A1A1A",
    "text_secondary":   "#6B7280",
    "text_label":       "#374151",

    "log_bg":           "#1A1A2E",
    "log_text":         "#D4D4D8",
    "log_error":        "#FCA5A5",
    "log_success":      "#86EFAC",

    "btn_run":          "#2D6A4F",
    "btn_run_hover":    "#40916C",
    "btn_stop":         "#DC2626",
    "btn_stop_hover":   "#EF4444",
    "btn_clear":        "#E5E7EB",
    "btn_clear_hover":  "#D1D5DB",

    "step_number_bg":   "#2D6A4F",
    "step_number_text": "#FFFFFF",

    "status_ok":        "#22C55E",
    "status_running":   "#FBBF24",
    "status_error":     "#EF4444",

    "switch_on":        "#2D6A4F",
    "switch_off":       "#9CA3AF",
}

# Step icons (unicode symbols)
STEP_ICONS = ["⬇", "✍", "📊", "🔗", "📄", "📨", "🗂️"]


class App:
    def __init__(self, root):
        self.root = root
        self.is_running = False

        # ── Window Configuration ──
        root.title("SIMAK-TPP - Dinas Kehutanan Provinsi Lampung")
        root.geometry("1100x750")
        root.minsize(950, 650)

        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("green")

        # ── Main Container ──
        main_container = ctk.CTkFrame(root, fg_color=COLORS["content_bg"], corner_radius=0)
        main_container.pack(fill="both", expand=True)
        main_container.grid_columnconfigure(1, weight=1)
        main_container.grid_rowconfigure(0, weight=1)

        # ══════════════════════════════════════════
        # SIDEBAR (LEFT)
        # ══════════════════════════════════════════
        sidebar = ctk.CTkFrame(main_container, width=230, fg_color=COLORS["sidebar_bg"], corner_radius=0)
        sidebar.grid(row=0, column=0, sticky="nsew")
        sidebar.grid_propagate(False)

        # ── Logo & Branding ──
        logo_frame = ctk.CTkFrame(sidebar, fg_color="transparent")
        logo_frame.pack(pady=(20, 5), padx=15)

        # Try to load the logo image
        app_dir = os.path.dirname(os.path.abspath(__file__))
        logo_path = os.path.join(app_dir, "files", "logo-pemprov.ico")

        if PIL_AVAILABLE and os.path.exists(logo_path):
            try:
                logo_image = ctk.CTkImage(
                    light_image=Image.open(logo_path),
                    dark_image=Image.open(logo_path),
                    size=(64, 64)
                )
                logo_label = ctk.CTkLabel(logo_frame, image=logo_image, text="")
                logo_label.pack()
            except Exception:
                # Fallback: show text placeholder
                logo_label = ctk.CTkLabel(
                    logo_frame, text="🏛️", font=ctk.CTkFont(size=40),
                    text_color=COLORS["sidebar_text"]
                )
                logo_label.pack()
        else:
            logo_label = ctk.CTkLabel(
                logo_frame, text="🏛️", font=ctk.CTkFont(size=40),
                text_color=COLORS["sidebar_text"]
            )
            logo_label.pack()

        ctk.CTkLabel(
            sidebar, text="DINAS KEHUTANAN",
            font=ctk.CTkFont(family="Helvetica", size=15, weight="bold"),
            text_color=COLORS["sidebar_text"]
        ).pack(pady=(5, 0))

        ctk.CTkLabel(
            sidebar, text="PROVINSI LAMPUNG",
            font=ctk.CTkFont(family="Helvetica", size=11),
            text_color=COLORS["sidebar_subtext"]
        ).pack(pady=(0, 15))

        # ── Separator ──
        separator = ctk.CTkFrame(sidebar, height=1, fg_color=COLORS["sidebar_hover"])
        separator.pack(fill="x", padx=15, pady=(0, 10))

        # ── Section Label & Master Checkbox ──
        section_header = ctk.CTkFrame(sidebar, fg_color="transparent")
        section_header.pack(fill="x", padx=15, pady=(0, 10))

        ctk.CTkLabel(
            section_header, text="SIMAK-TPP · PILIH PROSES",
            font=ctk.CTkFont(family="Helvetica", size=10),
            text_color=COLORS["sidebar_subtext"]
        ).pack(side="left", anchor="w")

        self.var_master_check = ctk.BooleanVar(value=True)
        self._master_check_updating = False  # Flag agar master toggle tidak trigger group logic

        def _on_master_check_toggle():
            val = self.var_master_check.get()
            self._master_check_updating = True
            for var in self.step_vars:
                var.set(val)
            self._master_check_updating = False
            self._update_form_visibility()

        self.master_checkbox = ctk.CTkCheckBox(
            section_header,
            text="Pilih Semua",
            variable=self.var_master_check,
            command=_on_master_check_toggle,
            font=ctk.CTkFont(family="Helvetica", size=10, weight="bold"),
            text_color=COLORS["sidebar_subtext"],
            fg_color=COLORS["switch_on"],
            hover_color=COLORS["accent_green_hover"],
            border_color=COLORS["sidebar_subtext"],
            checkbox_width=16,
            checkbox_height=16,
            corner_radius=4
        )
        self.master_checkbox.pack(side="right", anchor="e")

        # ── Process Steps ──
        self.step_definitions = [
            ("Download Rekap\nKehadiran",        "⬇"),
            ("Rekap Kehadiran\nApel",           "📋"),
            ("Analisis\nKehadiran",              "📊"),
            ("Mail Merge TPP",                  "📨"),
            ("Merge PDF\nper Jabatan",          "🗂️"),
            ("Pecah & Distribusi\nDokumen TTD", "✂️"),
        ]

        self.step_vars = []
        self.step_frames = []

        # Definisi grup eksklusif (harus ada SEBELUM switch dibuat):
        #   Grup A: Step 1 & 2 (index 0, 1)  — Download + Rekap Apel
        #   Grup B: Step 3-5   (index 2,3,4) — Analisis → Merge Jabatan
        #   Grup C: Step 6     (index 5)     — Pecah & Distribusi TTD
        self._step_groups = {
            "A": [0, 1],
            "B": [2, 3, 4],
            "C": [5],
        }
        self._index_to_group = {}
        for group_name, indices in self._step_groups.items():
            for idx in indices:
                self._index_to_group[idx] = group_name

        steps_container = ctk.CTkFrame(sidebar, fg_color="transparent")
        steps_container.pack(fill="both", expand=True, padx=10)

        for i, (label_text, icon) in enumerate(self.step_definitions):
            step_var = ctk.BooleanVar(value=True)
            self.step_vars.append(step_var)

            step_frame = ctk.CTkFrame(steps_container, fg_color="transparent", height=48)
            step_frame.pack(fill="x", pady=3)
            step_frame.pack_propagate(False)
            self.step_frames.append(step_frame)

            # Step number circle
            number_label = ctk.CTkLabel(
                step_frame, text=str(i + 1),
                width=28, height=28,
                corner_radius=14,
                fg_color=COLORS["step_number_bg"],
                text_color=COLORS["step_number_text"],
                font=ctk.CTkFont(size=13, weight="bold")
            )
            number_label.pack(side="left", padx=(5, 8))

            # Icon
            icon_label = ctk.CTkLabel(
                step_frame, text=icon,
                font=ctk.CTkFont(size=14),
                text_color=COLORS["sidebar_text"],
                width=20
            )
            icon_label.pack(side="left", padx=(0, 5))

            # Step name
            name_label = ctk.CTkLabel(
                step_frame, text=label_text,
                font=ctk.CTkFont(family="Helvetica", size=11),
                text_color=COLORS["sidebar_text"],
                anchor="w", justify="left"
            )
            name_label.pack(side="left", fill="x", expand=True)

            # Toggle switch — dengan callback untuk grup eksklusif
            step_index = i  # capture loop variable
            switch = ctk.CTkSwitch(
                step_frame, text="",
                variable=step_var,
                command=lambda idx=step_index: self._on_step_toggle(idx),
                width=42, height=22,
                switch_width=38, switch_height=19,
                progress_color=COLORS["switch_on"],
                fg_color=COLORS["switch_off"],
                button_color="#FFFFFF",
                button_hover_color="#F0F0F0"
            )
            switch.pack(side="right", padx=(0, 5))

        # Map step vars to variable names
        self.var_download      = self.step_vars[0]
        self.var_rekap_apel    = self.step_vars[1]
        self.var_analisis      = self.step_vars[2]
        self.var_mailmerge     = self.step_vars[3]
        self.var_merge_jabatan = self.step_vars[4]
        self.var_split_pdf     = self.step_vars[5]



        # ── Footer ──
        ctk.CTkLabel(
            sidebar, text="© 2026 Dinas Kehutanan Prov. Lampung",
            font=ctk.CTkFont(size=9),
            text_color=COLORS["sidebar_subtext"]
        ).pack(side="bottom", pady=10)

        # ══════════════════════════════════════════
        # CONTENT AREA (RIGHT)
        # ══════════════════════════════════════════
        content = ctk.CTkFrame(main_container, fg_color=COLORS["content_bg"], corner_radius=0)
        content.grid(row=0, column=1, sticky="nsew")
        content.grid_columnconfigure(0, weight=1)
        content.grid_rowconfigure(2, weight=1)  # Log console expands

        # ── Header Bar ──
        header = ctk.CTkFrame(content, height=50, fg_color=COLORS["card_bg"], corner_radius=0)
        header.grid(row=0, column=0, sticky="ew", padx=0, pady=0)
        header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            header,
            text="Sistem Informasi Manajemen Kehadiran & Tambahan Penghasilan Pegawai",
            font=ctk.CTkFont(family="Helvetica", size=14, weight="bold"),
            text_color=COLORS["text_primary"]
        ).grid(row=0, column=0, sticky="w", padx=20, pady=12)

        self.status_label = ctk.CTkLabel(
            header, text="● Siap dijalankan",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["status_ok"]
        )
        self.status_label.grid(row=0, column=1, sticky="e", padx=20, pady=12)

        # ── Form Area ──
        form_area = ctk.CTkFrame(content, fg_color="transparent")
        form_area.grid(row=1, column=0, sticky="ew", padx=20, pady=(15, 5))
        form_area.grid_columnconfigure(0, weight=3)
        form_area.grid_columnconfigure(1, weight=2)

        # ── Source Files Card ──
        source_card = ctk.CTkFrame(form_area, fg_color=COLORS["card_bg"], corner_radius=10)
        source_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)

        ctk.CTkLabel(
            source_card, text="📁  Source Files",
            font=ctk.CTkFont(family="Helvetica", size=13, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(anchor="w", padx=18, pady=(15, 10))

        self.entry_template_excel, self._frame_template_excel = self._create_file_input(
            source_card, "Template Excel", "xlsx", return_container=True
        )
        self.entry_data_pegawai, self._frame_data_pegawai = self._create_file_input(
            source_card, "Data Pegawai Excel *", "xlsx", return_container=True
        )
        self.entry_kalender_json, self._frame_kalender_json = self._create_file_input(
            source_card, "Kalender Kerja Excel (opsional)", "xlsx", return_container=True
        )
        self.entry_pdf_gabungan, self._frame_pdf_gabungan = self._create_file_input(
            source_card, "File PDF Gabungan (TTD) *", "pdf", return_container=True
        )

        # Bottom padding
        ctk.CTkFrame(source_card, height=10, fg_color="transparent").pack()

        # ── Right Column (Output + Period) ──
        right_col = ctk.CTkFrame(form_area, fg_color="transparent")
        right_col.grid(row=0, column=1, sticky="nsew", padx=0, pady=0)
        right_col.grid_columnconfigure(0, weight=1)

        # Output Settings Card
        output_card = ctk.CTkFrame(right_col, fg_color=COLORS["card_bg"], corner_radius=10)
        output_card.grid(row=0, column=0, sticky="nsew", padx=0, pady=(0, 10))

        ctk.CTkLabel(
            output_card, text="📂  Output Settings",
            font=ctk.CTkFont(family="Helvetica", size=13, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(anchor="w", padx=18, pady=(15, 10))

        self.entry_base_dir = self._create_folder_input(
            output_card, "Folder Penyimpanan Hasil"
        )

        # ── Download Worker Setting ──
        self.worker_section = ctk.CTkFrame(output_card, fg_color="transparent")
        self.worker_section.pack(fill="x", padx=18, pady=(0, 10))

        worker_header = ctk.CTkFrame(self.worker_section, fg_color="transparent")
        worker_header.pack(fill="x")

        ctk.CTkLabel(
            worker_header,
            text="⚡  Kecepatan Download",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_label"]
        ).pack(side="left", anchor="w")

        self.worker_value_label = ctk.CTkLabel(
            worker_header,
            text="5 worker  •  Normal",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=COLORS["accent_green"]
        )
        self.worker_value_label.pack(side="right", anchor="e")

        self.worker_var = tk.IntVar(value=5)

        def _on_worker_change(val):
            n = int(float(val))
            self.worker_var.set(n)
            if n <= 2:
                desc = "Lambat"
                color = "#6B7280"
            elif n <= 5:
                desc = "Normal"
                color = COLORS["accent_green"]
            elif n <= 7:
                desc = "Cepat"
                color = "#D97706"
            else:
                desc = "Turbo 🚀"
                color = "#DC2626"
            self.worker_value_label.configure(
                text=f"{n} worker  •  {desc}",
                text_color=color
            )

        worker_slider = ctk.CTkSlider(
            self.worker_section,
            from_=1, to=10,
            number_of_steps=9,
            variable=self.worker_var,
            command=_on_worker_change,
            progress_color=COLORS["accent_green"],
            button_color=COLORS["accent_green"],
            button_hover_color=COLORS["accent_green_hover"],
            fg_color=COLORS["card_border"],
            height=18
        )
        worker_slider.pack(fill="x", pady=(6, 2))

        hint_frame = ctk.CTkFrame(self.worker_section, fg_color="transparent")
        hint_frame.pack(fill="x")
        hint_frame.grid_columnconfigure(0, weight=0)
        hint_frame.grid_columnconfigure(1, weight=4)
        hint_frame.grid_columnconfigure(2, weight=0)
        hint_frame.grid_columnconfigure(3, weight=5)
        hint_frame.grid_columnconfigure(4, weight=0)

        ctk.CTkLabel(hint_frame, text="1", font=ctk.CTkFont(size=9), text_color=COLORS["text_secondary"]).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(hint_frame, text="5", font=ctk.CTkFont(size=9), text_color=COLORS["text_secondary"]).grid(row=0, column=2)
        ctk.CTkLabel(hint_frame, text="10", font=ctk.CTkFont(size=9), text_color=COLORS["text_secondary"]).grid(row=0, column=4, sticky="e")

        ctk.CTkLabel(
            self.worker_section,
            text="⚠ Worker lebih banyak = download lebih cepat, tapi lebih berat di RAM & CPU.",
            font=ctk.CTkFont(size=9),
            text_color=COLORS["text_secondary"],
            wraplength=220,
            justify="left"
        ).pack(anchor="w", pady=(4, 0))

        ctk.CTkFrame(output_card, height=4, fg_color="transparent").pack()

        # Period Card
        period_card = ctk.CTkFrame(right_col, fg_color=COLORS["card_bg"], corner_radius=10)
        period_card.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)

        ctk.CTkLabel(
            period_card, text="Period",
            font=ctk.CTkFont(family="Helvetica", size=13, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(anchor="w", padx=18, pady=(15, 5))

        ctk.CTkLabel(
            period_card, text="PERIODE",
            font=ctk.CTkFont(size=10),
            text_color=COLORS["text_secondary"]
        ).pack(anchor="w", padx=18, pady=(0, 8))

        period_inputs = ctk.CTkFrame(period_card, fg_color="transparent")
        period_inputs.pack(padx=18, pady=(0, 15), anchor="w")

        ctk.CTkLabel(
            period_inputs, text="Bulan",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_label"]
        ).grid(row=0, column=0, padx=(0, 8))

        self.entry_bulan = ctk.CTkEntry(
            period_inputs, width=55, height=32,
            font=ctk.CTkFont(size=13),
            border_color=COLORS["card_border"],
            fg_color=COLORS["card_bg"],
            text_color=COLORS["text_primary"],
            justify="center"
        )
        self.entry_bulan.grid(row=0, column=1, padx=(0, 20))

        ctk.CTkLabel(
            period_inputs, text="Tahun",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_label"]
        ).grid(row=0, column=2, padx=(0, 8))

        self.entry_tahun = ctk.CTkEntry(
            period_inputs, width=75, height=32,
            font=ctk.CTkFont(size=13),
            border_color=COLORS["card_border"],
            fg_color=COLORS["card_bg"],
            text_color=COLORS["text_primary"],
            justify="center"
        )
        self.entry_tahun.grid(row=0, column=3)

        # Pre-fill period
        now = datetime.now()
        bulan = now.month - 1 if now.month > 1 else 12
        self.entry_bulan.insert(0, f"{bulan:02d}")
        self.entry_tahun.insert(0, str(now.year))

        # ══════════════════════════════════════════
        # LOG CONSOLE (Tabbed — VS Code style)
        # ══════════════════════════════════════════
        self._log_entries = []  # Store all log entries: list of (timestamp, category, text)
        self._active_log_tab = "Semua"

        log_area = ctk.CTkFrame(content, fg_color="transparent")
        log_area.grid(row=2, column=0, sticky="nsew", padx=20, pady=(10, 0))
        log_area.grid_columnconfigure(0, weight=1)
        log_area.grid_rowconfigure(1, weight=1)

        # Log header with tabs
        log_header = ctk.CTkFrame(log_area, fg_color="transparent")
        log_header.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        log_header.grid_columnconfigure(1, weight=1)

        # Tab buttons container
        tab_container = ctk.CTkFrame(log_header, fg_color="transparent")
        tab_container.grid(row=0, column=0, sticky="w")

        self._log_tab_buttons = {}
        tab_names = ["Semua", "Error", "Warning", "Info"]
        tab_icons = {
            "Semua": ">_ ",
            "Error": "⊘ ",
            "Warning": "⚠ ",
            "Info": "ℹ ",
        }

        for i, tab_name in enumerate(tab_names):
            is_active = (tab_name == "Semua")
            btn = ctk.CTkButton(
                tab_container,
                text=f"{tab_icons[tab_name]}{tab_name}",
                width=85, height=28,
                font=ctk.CTkFont(size=11, weight="bold" if is_active else "normal"),
                fg_color=COLORS["log_bg"] if is_active else "transparent",
                hover_color="#2A2A40",
                text_color="#FFFFFF" if is_active else COLORS["text_secondary"],
                corner_radius=6,
                command=lambda t=tab_name: self._switch_log_tab(t)
            )
            btn.grid(row=0, column=i, padx=(0, 4))
            self._log_tab_buttons[tab_name] = btn

        # Error/Warning counter badges
        self._error_count = 0
        self._warning_count = 0

        # Clear button
        btn_clear = ctk.CTkButton(
            log_header, text="Bersihkan", width=90, height=28,
            font=ctk.CTkFont(size=11),
            fg_color=COLORS["btn_clear"],
            hover_color=COLORS["btn_clear_hover"],
            text_color=COLORS["text_primary"],
            corner_radius=6,
            command=self._clear_log
        )
        btn_clear.grid(row=0, column=1, sticky="e")

        # Log textbox (terminal style)
        self.log_box = ctk.CTkTextbox(
            log_area, height=180,
            fg_color=COLORS["log_bg"],
            text_color=COLORS["log_text"],
            font=ctk.CTkFont(family="Courier New", size=12),
            corner_radius=8,
            border_width=0,
            scrollbar_button_color=COLORS["sidebar_hover"],
            scrollbar_button_hover_color=COLORS["accent_green"],
            wrap="word"
        )
        self.log_box.grid(row=1, column=0, sticky="nsew")

        # Configure text tags for colored log lines
        self.log_box._textbox.tag_configure("error", foreground="#FCA5A5")
        self.log_box._textbox.tag_configure("warning", foreground="#FCD34D")
        self.log_box._textbox.tag_configure("success", foreground="#86EFAC")
        self.log_box._textbox.tag_configure("info", foreground="#D4D4D8")
        self.log_box._textbox.tag_configure("step", foreground="#93C5FD")

        # ── Action Button Area ──
        action_area = ctk.CTkFrame(content, fg_color="transparent", height=60)
        action_area.grid(row=3, column=0, sticky="ew", padx=20, pady=(10, 15))
        action_area.grid_columnconfigure(0, weight=1)

        self.btn_action = ctk.CTkButton(
            action_area,
            text="▶   JALANKAN",
            width=180, height=45,
            font=ctk.CTkFont(family="Helvetica", size=15, weight="bold"),
            fg_color=COLORS["btn_run"],
            hover_color=COLORS["btn_run_hover"],
            text_color="#FFFFFF",
            corner_radius=10,
            command=self.toggle_process
        )
        self.btn_action.grid(row=0, column=0, sticky="e")

        # ── Firebase Access Validator ──
        self.access_validator = ValidateUserAccess(
            gateway=FirebaseRealtimeUserAccessGateway(base_url=FIREBASE_URL)
        )

    # ══════════════════════════════════════════════════════════
    # Helper: Create a file input row
    # ══════════════════════════════════════════════════════════
    def _create_file_input(self, parent, label_text, file_type, return_container=False):
        """Create a labeled file input with a browse button and drag-and-drop."""
        container = ctk.CTkFrame(parent, fg_color="transparent")
        container.pack(fill="x", padx=18, pady=(0, 8))

        ctk.CTkLabel(
            container, text=label_text,
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_label"]
        ).pack(anchor="w", pady=(0, 4))

        row = ctk.CTkFrame(container, fg_color="transparent")
        row.pack(fill="x")
        row.grid_columnconfigure(0, weight=1)

        # Use a StringVar to hold the path value
        path_var = tk.StringVar()

        placeholder = f"Drag & drop file {file_type.upper()} atau klik untuk browse"

        entry = ctk.CTkEntry(
            row, height=35,
            font=ctk.CTkFont(size=11),
            border_color=COLORS["card_border"],
            fg_color="#F9FAFB",
            text_color=COLORS["text_primary"],
            placeholder_text=placeholder,
            placeholder_text_color="#A0AEC0",
            corner_radius=6,
            textvariable=path_var,
            state="readonly"
        )
        entry.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        if file_type == "xlsx":
            filetypes = [("Excel Workbook", "*.xlsx")]
            btn_icon = "📗"
        elif file_type == "pdf":
            filetypes = [("PDF File", "*.pdf")]
            btn_icon = "📕"
        else:
            filetypes = [("JSON File", "*.json")]
            btn_icon = "📋"

        def browse(event=None):
            path = filedialog.askopenfilename(
                title=f"Pilih File {file_type.upper()}",
                filetypes=filetypes
            )
            if path:
                path_var.set(path)

        # Clicking anywhere on the entry opens the file dialog
        entry.bind("<Button-1>", browse)

        # Register drag-and-drop target
        if DND_AVAILABLE:
            inner = entry._entry if hasattr(entry, '_entry') else entry
            try:
                inner.drop_target_register(DND_FILES)
                def on_drop(event):
                    # Clean path from braces (macOS sometimes wraps paths)
                    dropped = event.data.strip().strip('{}')
                    ext = os.path.splitext(dropped)[1].lower()
                    valid_exts = [f'.{file_type}']
                    
                    if ext in valid_exts:
                        path_var.set(dropped)
                    else:
                        messagebox.showwarning("Format Salah", f"File harus berformat {', '.join(valid_exts)}")
                    entry.configure(border_color=COLORS["card_border"])
                def on_enter(event):
                    entry.configure(border_color=COLORS["accent_green"])
                    return event.action
                def on_leave(event):
                    entry.configure(border_color=COLORS["card_border"])
                inner.dnd_bind('<<Drop>>', on_drop)
                inner.dnd_bind('<<DragEnter>>', on_enter)
                inner.dnd_bind('<<DragLeave>>', on_leave)
            except Exception:
                pass  # DnD registration failed, click-to-browse still works

        # Clear button (✕)
        btn_clear = ctk.CTkButton(
            row, text="✕", width=30, height=35,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color="transparent",
            hover_color="#FEE2E2",
            text_color="#9CA3AF",
            corner_radius=6,
            command=lambda: path_var.set("")
        )
        btn_clear.grid(row=0, column=1, padx=(0, 4))

        btn = ctk.CTkButton(
            row, text=btn_icon, width=40, height=35,
            font=ctk.CTkFont(size=16),
            fg_color=COLORS["accent_green"],
            hover_color=COLORS["accent_green_hover"],
            text_color="#FFFFFF",
            corner_radius=6,
            command=browse
        )
        btn.grid(row=0, column=2)

        # Store path_var on the entry for .get() compatibility
        entry._path_var = path_var

        if return_container:
            return entry, container
        return entry

    def _create_folder_input(self, parent, label_text):
        """Create a labeled folder input with a browse button and drag-and-drop."""
        container = ctk.CTkFrame(parent, fg_color="transparent")
        container.pack(fill="x", padx=18, pady=(0, 8))

        ctk.CTkLabel(
            container, text=label_text,
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_label"]
        ).pack(anchor="w", pady=(0, 4))

        row = ctk.CTkFrame(container, fg_color="transparent")
        row.pack(fill="x")
        row.grid_columnconfigure(0, weight=1)

        # Use a StringVar to hold the path value
        path_var = tk.StringVar()

        entry = ctk.CTkEntry(
            row, height=35,
            font=ctk.CTkFont(size=11),
            border_color=COLORS["card_border"],
            fg_color="#F9FAFB",
            text_color=COLORS["text_primary"],
            placeholder_text="Drag & drop folder atau klik untuk browse",
            placeholder_text_color="#A0AEC0",
            corner_radius=6,
            textvariable=path_var,
            state="readonly"
        )
        entry.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        def browse(event=None):
            path = filedialog.askdirectory(title="Pilih Folder Penyimpanan")
            if path:
                path_var.set(path)

        # Clicking anywhere on the entry opens the folder dialog
        entry.bind("<Button-1>", browse)

        # Register drag-and-drop target
        if DND_AVAILABLE:
            inner = entry._entry if hasattr(entry, '_entry') else entry
            try:
                inner.drop_target_register(DND_FILES)
                def on_drop(event):
                    dropped = event.data.strip().strip('{}')
                    if os.path.isdir(dropped):
                        path_var.set(dropped)
                    else:
                        # If a file is dropped, use its parent directory
                        parent_dir = os.path.dirname(dropped)
                        if os.path.isdir(parent_dir):
                            path_var.set(parent_dir)
                    entry.configure(border_color=COLORS["card_border"])
                def on_enter(event):
                    entry.configure(border_color=COLORS["accent_green"])
                    return event.action
                def on_leave(event):
                    entry.configure(border_color=COLORS["card_border"])
                inner.dnd_bind('<<Drop>>', on_drop)
                inner.dnd_bind('<<DragEnter>>', on_enter)
                inner.dnd_bind('<<DragLeave>>', on_leave)
            except Exception:
                pass  # DnD registration failed, click-to-browse still works

        # Clear button (✕)
        btn_clear = ctk.CTkButton(
            row, text="✕", width=30, height=35,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color="transparent",
            hover_color="#FEE2E2",
            text_color="#9CA3AF",
            corner_radius=6,
            command=lambda: path_var.set("")
        )
        btn_clear.grid(row=0, column=1, padx=(0, 4))

        btn = ctk.CTkButton(
            row, text="📁", width=40, height=35,
            font=ctk.CTkFont(size=16),
            fg_color=COLORS["accent_green"],
            hover_color=COLORS["accent_green_hover"],
            text_color="#FFFFFF",
            corner_radius=6,
            command=browse
        )
        btn.grid(row=0, column=2)

        # Store path_var on the entry for .get() compatibility
        entry._path_var = path_var

        return entry

    # ══════════════════════════════════════════════════════════
    # Log functions (tabbed, color-coded)
    # ══════════════════════════════════════════════════════════
    def _categorize(self, text):
        """Determine the category of a log message."""
        t = str(text).lower()
        if any(k in t for k in ["error", "gagal", "❌", "traceback", "exception", "kesalahan"]):
            return "error"
        if any(k in t for k in ["warning", "⚠", "peringatan"]):
            return "warning"
        if any(k in t for k in ["✅", "✔", "selesai", "sukses", "berhasil"]):
            return "success"
        if any(k in t for k in ["──", "==", "---"]):
            return "step"
        return "info"

    def log(self, text):
        """Thread-safe logging to the console textbox with category tagging."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        category = self._categorize(text)
        entry = (timestamp, category, str(text))
        self._log_entries.append(entry)

        # Update counters
        if category == "error":
            self._error_count += 1
        elif category == "warning":
            self._warning_count += 1

        def _update():
            # Update tab button labels with counts
            if self._error_count > 0:
                self._log_tab_buttons["Error"].configure(
                    text=f"⊘ Error ({self._error_count})"
                )
            if self._warning_count > 0:
                self._log_tab_buttons["Warning"].configure(
                    text=f"⚠ Warning ({self._warning_count})"
                )

            # Only display if matching current tab
            if self._active_log_tab == "Semua" or \
               (self._active_log_tab == "Error" and category == "error") or \
               (self._active_log_tab == "Warning" and category == "warning") or \
               (self._active_log_tab == "Info" and category in ("info", "step", "success")):
                line = f"[{timestamp}] {text}\n"
                self.log_box._textbox.configure(state="normal")
                self.log_box._textbox.insert("end", line, category)
                self.log_box._textbox.configure(state="disabled")
                self.log_box.see("end")

        self.root.after(0, _update)

    def _switch_log_tab(self, tab_name):
        """Switch the active log tab and re-render filtered messages."""
        self._active_log_tab = tab_name

        # Update tab button styles
        for name, btn in self._log_tab_buttons.items():
            if name == tab_name:
                btn.configure(
                    fg_color=COLORS["log_bg"],
                    text_color="#FFFFFF",
                    font=ctk.CTkFont(size=11, weight="bold")
                )
            else:
                btn.configure(
                    fg_color="transparent",
                    text_color=COLORS["text_secondary"],
                    font=ctk.CTkFont(size=11)
                )

        # Re-render log with filter
        self.log_box._textbox.configure(state="normal")
        self.log_box._textbox.delete("1.0", "end")

        for ts, cat, text in self._log_entries:
            show = False
            if tab_name == "Semua":
                show = True
            elif tab_name == "Error" and cat == "error":
                show = True
            elif tab_name == "Warning" and cat == "warning":
                show = True
            elif tab_name == "Info" and cat in ("info", "step", "success"):
                show = True

            if show:
                line = f"[{ts}] {text}\n"
                self.log_box._textbox.insert("end", line, cat)

        self.log_box._textbox.configure(state="disabled")
        self.log_box.see("end")

    def _clear_log(self):
        """Clear the log console and reset counters."""
        self._log_entries.clear()
        self._error_count = 0
        self._warning_count = 0
        self.log_box._textbox.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box._textbox.configure(state="disabled")

        # Reset tab labels
        self._log_tab_buttons["Error"].configure(text="⊘ Error")
        self._log_tab_buttons["Warning"].configure(text="⚠ Warning")

    def _on_step_toggle(self, toggled_index):
        """Enforce mutual exclusivity antara 3 grup step.
        Grup A: Step 1 & 2  (Download + Rekap Apel)
        Grup B: Step 3-5    (Analisis → Merge Jabatan)
        Grup C: Step 6      (Pecah & Distribusi TTD)
        Ketika step di satu grup dinyalakan, semua step di grup lain otomatis mati.
        """
        # Safety guard: jangan jalankan jika init belum selesai
        if not hasattr(self, '_step_groups') or not hasattr(self, '_master_check_updating'):
            return
        # Jangan jalankan logika grup saat master checkbox toggle semua sekaligus
        if self._master_check_updating:
            return

        is_on = self.step_vars[toggled_index].get()
        if not is_on:
            # Jika user mematikan step, tidak perlu matikan grup lain
            return

        my_group = self._index_to_group.get(toggled_index)
        if not my_group:
            return

        # Nyalakan SEMUA step di grup yang sama
        for idx in self._step_groups[my_group]:
            self.step_vars[idx].set(True)

        # Matikan semua step di grup lain
        for group_name, indices in self._step_groups.items():
            if group_name != my_group:
                for idx in indices:
                    self.step_vars[idx].set(False)

        # Update master checkbox: centang hanya jika SEMUA step hidup
        all_on = all(var.get() for var in self.step_vars)
        self.var_master_check.set(all_on)

        # Update tampilan form sesuai grup aktif
        self._update_form_visibility()

    def _update_form_visibility(self):
        """Show/hide form fields berdasarkan grup step yang aktif.
        Grup A (Step 1-2): Data Pegawai, Ekin & Apel, Worker Slider
        Grup B (Step 3-5): Template Excel, Ekin & Apel, Kalender Kerja
        Grup C (Step 6):   PDF Gabungan
        Folder Hasil & Periode selalu tampil.
        """
        # Safety guard
        if not hasattr(self, '_frame_template_excel'):
            return

        # Tentukan grup mana yang aktif berdasarkan step yang hidup
        active_groups = []
        for group_name, indices in self._step_groups.items():
            if any(self.step_vars[idx].get() for idx in indices):
                active_groups.append(group_name)

        # Jika lebih dari 1 grup aktif (Pilih Semua) atau tidak ada step aktif → tampilkan semua
        if len(active_groups) != 1:
            active_group = "ALL"
        else:
            active_group = active_groups[0]

        # Mapping: field frame → grup mana saja yang membutuhkannya
        field_visibility = {
            self._frame_data_pegawai:    ["A", "ALL"],
            self._frame_template_excel:  ["B", "ALL"],
            self._frame_kalender_json:   ["B", "ALL"],
            self._frame_pdf_gabungan:    ["C", "ALL"],
        }

        pack_opts = {"fill": "x", "padx": 18, "pady": (0, 8)}

        for frame, groups in field_visibility.items():
            if active_group in groups:
                if not frame.winfo_ismapped():
                    frame.pack(**pack_opts)
            else:
                if frame.winfo_ismapped():
                    frame.pack_forget()

        # Worker slider: hanya tampil untuk Grup A (Download)
        worker_pack_opts = {"fill": "x", "padx": 18, "pady": (0, 10)}
        if active_group in ["A", "ALL"]:
            if not self.worker_section.winfo_ismapped():
                self.worker_section.pack(**worker_pack_opts)
        else:
            if self.worker_section.winfo_ismapped():
                self.worker_section.pack_forget()

    def _set_status(self, text, color):
        """Update the status label in the header."""
        self.status_label.configure(text=text, text_color=color)

    # ══════════════════════════════════════════════════════════
    # Business Logic (preserved 100% from original)
    # ══════════════════════════════════════════════════════════
    def jalankan(self):
        try:
            raw_base_dir = self.entry_base_dir.get().strip() if self.entry_base_dir.get() else ""
            if not raw_base_dir or not os.path.exists(raw_base_dir):
                self.log("❌ Folder Utama / Folder Hasil belum dipilih atau tidak ditemukan.")
                self.root.after(0, lambda: self._set_status("● Folder Utama belum dipilih", COLORS["status_error"]))
                self.root.after(0, lambda: messagebox.showerror(
                    "Peringatan Input",
                    "Folder Utama / Folder Hasil belum dipilih atau folder tidak ditemukan!\n\nHarap pilih Folder Hasil pada bagian kanan atas terlebih dahulu."
                ))
                return

            base_dir = os.path.abspath(raw_base_dir)
            base_dir = os.path.normpath(base_dir)
            base_dir = base_dir.strip()

            # Validasi minimal 1 step diaktifkan
            any_step = any([
                self.var_download.get(),
                self.var_rekap_apel.get(),
                self.var_analisis.get(),
                self.var_mailmerge.get(),
                self.var_merge_jabatan.get(),
                self.var_split_pdf.get()
            ])

            if not any_step:
                self.log("⚠️ Tidak ada langkah kerja (step) yang diaktifkan.")
                self.root.after(0, lambda: self._set_status("● Pilih minimal 1 step", COLORS["status_error"]))
                self.root.after(0, lambda: messagebox.showwarning(
                    "Peringatan Input",
                    "Tidak ada langkah kerja yang diaktifkan!\n\nHarap aktifkan minimal 1 saklar step di sebelah kiri sebelum menekan tombol JALANKAN."
                ))
                return

            # Buat struktur folder
            DIR_REKAP = os.path.join(base_dir, "REKAP KEHADIRAN")
            DIR_REKAP_DITANDATANGANI = os.path.join(base_dir, "REKAP KEHADIRAN DITANDATANGANI")
            DIR_OUTPUT = os.path.join(base_dir, "PERHITUNGAN TPP")
            TEMP_DIR = os.path.join(base_dir, "TEMP")

            DIR_REKAP_DITANDATANGANI = os.path.abspath(DIR_REKAP_DITANDATANGANI)
            DIR_OUTPUT = os.path.abspath(DIR_OUTPUT)
            TEMP_DIR = os.path.abspath(TEMP_DIR)

            os.makedirs(DIR_REKAP_DITANDATANGANI, exist_ok=True)
            os.makedirs(DIR_OUTPUT, exist_ok=True)
            os.makedirs(TEMP_DIR, exist_ok=True)

            excel = self.entry_template_excel.get()
            word = TEMPLATE_WORD
            ekin_apel = None
            excel_pegawai = self.entry_data_pegawai.get()
            json_kalender = self.entry_kalender_json.get()

            bulan = self.entry_bulan.get().zfill(2)
            tahun = int(self.entry_tahun.get())

            output_excel = os.path.join(base_dir, "Disiplin_TPP.xlsx")
            output_template_ready = os.path.join(base_dir, "Disiplin_TPP_Lengkap.xlsx")
            csv_output = os.path.join(base_dir, "Disiplin_TPP_Lengkap.csv")

            # Helper untuk menampilkan error log & messagebox thread-safe
            def show_err(msg, title="Peringatan Input"):
                self.log(f"❌ {msg}")
                self.root.after(0, lambda: self._set_status("● Error terjadi", COLORS["status_error"]))
                self.root.after(0, lambda: messagebox.showerror(title, msg))

            # STEP 1 — DOWNLOAD + TTD OTOMATIS
            if self.var_download.get():
                if stop_event.is_set():
                    return
                if not excel_pegawai or not os.path.exists(excel_pegawai):
                    show_err("File Data Pegawai Excel belum dipilih atau file tidak ditemukan.")
                    return
                self.log("── [1/6] Download Rekap Kehadiran ──")
                max_workers = self.worker_var.get()
                download_rekap(excel_pegawai, DIR_REKAP, bulan, tahun, self.log, max_workers)

                # Otomatis Input Spesimen TTD
                if stop_event.is_set():
                    return
                self.log("  📝 Input Spesimen Tanda Tangan (otomatis)...")
                input_spesimen(DIR_REKAP, DIR_REKAP_DITANDATANGANI, self.log)

                # Hapus folder REKAP KEHADIRAN (asli) karena sudah ada versi TTD
                import shutil
                import stat
                import time

                def _remove_readonly(func, path, exc_info):
                    try:
                        os.chmod(path, stat.S_IWRITE)
                        func(path)
                    except Exception:
                        pass

                if os.path.exists(DIR_REKAP):
                    try:
                        shutil.rmtree(DIR_REKAP, onerror=_remove_readonly)
                        self.log(f"  🗑 Folder REKAP KEHADIRAN dihapus (menyisakan hasil TTD).")
                    except Exception:
                        time.sleep(0.5)
                        try:
                            shutil.rmtree(DIR_REKAP, ignore_errors=True)
                            self.log(f"  🗑 Folder REKAP KEHADIRAN dihapus (menyisakan hasil TTD).")
                        except Exception:
                            self.log(f"  ⚠️ Folder REKAP KEHADIRAN mentah belum dapat dihapus penuh oleh Windows/OneDrive, namun proses tetap dilanjutkan.")

            # STEP 2 — REKAP KEHADIRAN APEL
            if self.var_rekap_apel.get():
                if stop_event.is_set():
                    return
                if not excel_pegawai or not os.path.exists(excel_pegawai):
                    show_err("File Data Pegawai Excel belum dipilih atau file tidak ditemukan.")
                    return
                self.log("── [2/6] Rekap Kehadiran Apel ──")
                rekap_kehadiran_apel(DIR_REKAP_DITANDATANGANI, excel_pegawai, bulan, tahun, None, self.log, ekin_apel)

            # STEP 3 — ANALISIS KEHADIRAN
            if self.var_analisis.get():
                if stop_event.is_set():
                    return
                if not excel or not os.path.exists(excel):
                    show_err("Template Excel belum dipilih atau file tidak ditemukan.")
                    return
                self.log("── [3/6] Analisis Kehadiran ──")
                analisis_kehadiran(DIR_REKAP_DITANDATANGANI, excel, output_excel, self.log, json_kalender)

            # GABUNG EKIN & APEL (OTOMATIS TANPA TOGGLE)
            # Selalu dijalankan jika Analisis Kehadiran aktif atau output_excel ada
            if self.var_analisis.get() or os.path.exists(output_excel):
                if stop_event.is_set():
                    return
                self.log("  🔗 Menggabungkan Data Ekin & Apel (otomatis)...")

                from utils import sanitize_filename
                path_ekin_terisi = os.path.join(DIR_REKAP_DITANDATANGANI, sanitize_filename(f"Ekin_Apel_Terisi_{bulan}_{tahun}.xlsx"))

                if not os.path.exists(path_ekin_terisi) and excel_pegawai and os.path.exists(excel_pegawai):
                    try:
                        from steps.rekap_apel import _baca_data_pegawai
                        pegawai_list, _ = _baca_data_pegawai(excel_pegawai, self.log)
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, Alignment, Border, Side

                        wb_new = Workbook()
                        ws_new = wb_new.active
                        ws_new.title = "DISIPLIN"

                        headers_template = ["No.", "Nama Pegawai", "NIP", "Predikat Kinerja", "TMA", "TMA Lain"]
                        ws_new.append(headers_template)

                        thin_border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thin', color='000000')
                        )

                        for col_idx in range(1, 7):
                            cell = ws_new.cell(row=1, column=col_idx)
                            cell.font = Font(name="Calibri", size=11, bold=True)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border = thin_border

                        seen_nips = set()
                        no_counter = 1
                        for p in pegawai_list:
                            nip_raw = str(p.get("nip", "")).strip()
                            nip_clean = nip_raw.replace(" ", "").strip()
                            if not nip_clean or nip_clean.lower() in ["nan", "none", "-", "0"]:
                                continue
                            if nip_clean in seen_nips:
                                continue
                            seen_nips.add(nip_clean)
                            nama_val = p.get("nama") or p.get("name") or ""
                            ws_new.append([no_counter, nama_val, nip_raw, "Baik/Sangat Baik", 0, 0])

                            row_idx = ws_new.max_row
                            for col_idx in range(1, 7):
                                c = ws_new.cell(row=row_idx, column=col_idx)
                                c.font = Font(name="Calibri", size=11)
                                c.border = thin_border
                                if col_idx in [1, 5, 6]:
                                    c.alignment = Alignment(horizontal="center", vertical="center")
                                else:
                                    c.alignment = Alignment(horizontal="left", vertical="center")
                            no_counter += 1

                        for col in ws_new.columns:
                            max_len = 0
                            col_letter = col[0].column_letter
                            for cell in col:
                                val_str = str(cell.value or "")
                                if len(val_str) > max_len:
                                    max_len = len(val_str)
                            ws_new.column_dimensions[col_letter].width = max(max_len + 4, 12)

                        wb_new.save(path_ekin_terisi)
                        wb_new.close()
                    except Exception:
                        pass

                if os.path.exists(path_ekin_terisi):
                    status, pesan = merge_ekin_apel(
                        path_ekin_terisi,
                        output_excel,
                        output_template_ready
                    )
                    self.log(f"  {pesan}")
                    if not status:
                        show_err(pesan, title="Error Gabung Ekin & Apel")
                        return
                else:
                    import shutil
                    shutil.copy(output_excel, output_template_ready)

            # STEP 4 — MAIL MERGE TPP (Termasuk Generate CSV Otomatis)
            if self.var_mailmerge.get():
                if stop_event.is_set():
                    return
                if not word or not os.path.exists(word):
                    show_err("Template Word TPP tidak ditemukan.")
                    return
                if not os.path.exists(output_template_ready):
                    show_err("File Disiplin_TPP_Lengkap.xlsx tidak ditemukan di Folder Hasil. Jalankan step Analisis Kehadiran (Step 3) terlebih dahulu.", title="Prasyarat Belum Ada")
                    return

                # Otomatis Generate CSV dari Disiplin_TPP_Lengkap.xlsx
                self.log("  📄 Generating CSV (otomatis)...")
                excel_sheet_disiplin_ke_csv(output_template_ready, base_dir)

                self.log("── [4/6] Mail Merge TPP ──")
                process_mail_merge(DIR_REKAP_DITANDATANGANI, DIR_OUTPUT, TEMP_DIR, csv_output, word, self.log)

            # STEP 5 — MERGE PDF PER JABATAN
            if self.var_merge_jabatan.get():
                if stop_event.is_set():
                    return
                if not os.path.exists(csv_output):
                    if os.path.exists(output_template_ready):
                        self.log("  📄 Generating CSV (otomatis)...")
                        excel_sheet_disiplin_ke_csv(output_template_ready, base_dir)
                    else:
                        show_err("File CSV (Disiplin_TPP_Lengkap.csv) / Disiplin_TPP_Lengkap.xlsx tidak ditemukan. Jalankan step sebelumnya terlebih dahulu.", title="Prasyarat Belum Ada")
                        return
                self.log("── [5/6] Merge PDF per Jabatan ──")
                merge_pdf_by_jabatan(DIR_OUTPUT, csv_output, bulan, tahun, self.log)

            # STEP 6 — PECAH & DISTRIBUSI DOKUMEN TTD
            if self.var_split_pdf.get():
                if stop_event.is_set():
                    return
                
                pdf_gabungan = self.entry_pdf_gabungan.get().strip()
                
                # 1. Wajibkan memasukkan File PDF Gabungan (TTD) hasil scan fisik
                if not pdf_gabungan or not os.path.exists(pdf_gabungan):
                    if not pdf_gabungan:
                        show_err("File PDF Gabungan (TTD) belum dipilih!\n\nHarap masukkan file PDF Gabungan (TTD) hasil scan fisik pada kolom di layar sebelum menjalankan step ini.")
                    else:
                        show_err(f"File PDF Gabungan (TTD) yang dipilih tidak ditemukan:\n{pdf_gabungan}")
                    return

                # 2. Cek ketersediaan file CSV pendukung data pegawai
                if not os.path.exists(csv_output):
                    if os.path.exists(output_template_ready):
                        self.log("  📄 Generating CSV (otomatis)...")
                        excel_sheet_disiplin_ke_csv(output_template_ready, base_dir)
                    else:
                        show_err("File CSV (Disiplin_TPP_Lengkap.csv) tidak ditemukan di Folder Hasil.\n\nFile CSV ini diperlukan untuk memetakan halaman PDF ke masing-masing pegawai. Jalankan Step 3 & 4 terlebih dahulu.", title="Prasyarat Belum Ada")
                        return

                self.log("── [6/6] Pecah & Distribusi Dokumen TTD ──")
                split_pdf_jabatan(DIR_OUTPUT, csv_output, bulan, tahun, self.log, pdf_gabungan)

            # Jika berhasil semua
            if not stop_event.is_set():
                self.log("✅ Semua proses selesai dengan sukses.")
                self.root.after(0, lambda: self._set_status("● Semua proses selesai!", COLORS["status_ok"]))
                self.root.after(0, lambda: messagebox.showinfo("Selesai", "Semua proses selesai"))

        # BAGIAN INI AKAN MENANGKAP ERROR APAPUN YANG BIKIN STUCK
        except SystemExit:
            # Jika diberhentikan paksa oleh pengguna
            pass
        except Exception as e:
            error_detail = traceback.format_exc() # Mengambil detail baris yang error
            self.log("\n================ ERROR TERJADI ================")
            self.log(error_detail)
            self.log("===============================================")

            self.root.after(0, lambda: self._set_status("● Error terjadi", COLORS["status_error"]))

            # Memunculkan pop-up error ke layar pengguna
            error_msg = f"Aplikasi berhenti karena error:\n{str(e)}\n\nSilakan cek kotak log di aplikasi untuk detailnya."
            self.root.after(0, lambda: messagebox.showerror("Terjadi Kesalahan Kritis", error_msg))
        finally:
            self.is_running = False
            # Pastikan seluruh browser Selenium/Chrome ditutup dan tidak ada proses worker yang menggantung di Task Manager
            try:
                from steps.download import cleanup_all_drivers
                cleanup_all_drivers()
            except Exception:
                pass

            self.root.after(0, lambda: self.btn_action.configure(
                text="▶   JALANKAN",
                fg_color=COLORS["btn_run"],
                hover_color=COLORS["btn_run_hover"],
                state="normal"
            ))
            if stop_event.is_set():
                self.log("\n================ PROSES DIBERHENTIKAN ================")
                self.log("Proses telah diberhentikan oleh pengguna.")
                self.root.after(0, lambda: messagebox.showinfo("Berhenti", "Proses berhasil diberhentikan. Silahkan mulai lagi."))
            
            # Reset status if still showing "running" (e.g. after early return or stop)
            def _reset_status_if_needed():
                current = self.status_label.cget("text")
                if stop_event.is_set():
                    self._set_status("● Proses diberhentikan", COLORS["status_running"])
                elif "berjalan" in current or "Menghentikan" in current:
                    self._set_status("● Siap dijalankan", COLORS["status_ok"])
            self.root.after(100, _reset_status_if_needed)

    def toggle_process(self):
        if getattr(self, 'is_running', False):
            self.stop_process()
        else:
            self.start_process()

    def start_process(self):
        is_allowed = self.access_validator.execute(APP_USERNAME)
        if is_allowed is None:
            self.root.after(0, lambda: messagebox.showerror(
                "Tidak Ada Koneksi",
                "Gagal terhubung ke server.\n\n"
                "Pastikan laptop terhubung ke internet, lalu coba jalankan kembali."
            ))
            return
        if not is_allowed:
            self.root.after(0, lambda: messagebox.showerror(
                "Akses Ditolak",
                "Akun ini tidak memiliki izin untuk menjalankan aplikasi.\n\n"
                "Hubungi administrator untuk mendapatkan akses."
            ))
            return

        stop_event.clear()
        self.is_running = True
        self.btn_action.configure(
            text="■   STOP",
            fg_color=COLORS["btn_stop"],
            hover_color=COLORS["btn_stop_hover"]
        )
        self._set_status("● Proses sedang berjalan...", COLORS["status_running"])
        self._clear_log()
        self.process_thread = threading.Thread(target=self.jalankan, daemon=True)
        self.process_thread.start()

    def stop_process(self):
        self.btn_action.configure(
            text="MENGHENTIKAN...",
            fg_color="orange",
            hover_color="orange",
            state="disabled"
        )
        self._set_status("● Menghentikan...", COLORS["status_running"])
        self.log("\n================ MENGHENTIKAN PROSES ================")
        self.log("Sedang menghentikan proses, mohon tunggu...")
        
        # Stop download and trigger cleanup
        threading.Thread(target=stop_download, daemon=True).start()
