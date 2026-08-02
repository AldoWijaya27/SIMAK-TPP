import json
import os
import sys
import time
import base64
import threading
import subprocess
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

from selenium import webdriver
from selenium.webdriver.chrome.webdriver import WebDriver as ChromeDriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

from utils import sanitize_filename

thread_local = threading.local()
drivers = []
drivers_lock = threading.Lock()
shared_cookies = []
stop_event = threading.Event()
current_log = None

def cleanup_driver(d):
    """Menutup driver Selenium beserta seluruh proses Chrome yang terkait."""
    if d is None:
        return

    # Ambil PID chromedriver sebelum quit, agar bisa di-taskkill jika quit hang
    pid = None
    try:
        if hasattr(d, "service") and hasattr(d.service, "process") and d.service.process:
            pid = d.service.process.pid
    except Exception:
        pass

    # Jalankan d.quit() di thread terpisah dengan timeout 5 detik
    # Ini mencegah hang lama di laptop dengan spesifikasi rendah / RAM terbatas
    quit_done = threading.Event()

    def _quit():
        try:
            d.quit()
        except Exception:
            pass
        finally:
            quit_done.set()

    t = threading.Thread(target=_quit, daemon=True)
    t.start()
    t.join(timeout=5.0)  # Maksimal tunggu 5 detik

    # Jika d.quit() tidak membersihkan child process di Windows, bunuh pohon prosesnya (process tree)
    try:
        if pid and sys.platform == "win32":
            subprocess.run(
                ["taskkill", "/F", "/T", "/PID", str(pid)],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=0x08000000  # CREATE_NO_WINDOW
            )
        elif pid:
            try:
                d.service.process.kill()
            except Exception:
                pass
    except Exception:
        pass

def cleanup_all_drivers():
    """Menutup seluruh browser Selenium yang aktif dan membebaskan memori."""
    with drivers_lock:
        active_drivers = list(drivers)
        drivers.clear()

    threads = []
    for driver in active_drivers:
        t = threading.Thread(target=cleanup_driver, args=(driver,), daemon=True)
        t.start()
        threads.append(t)

    for t in threads:
        t.join(timeout=3.0)

def stop_download():
    stop_event.set()
    if current_log:
        try:
            current_log("\nMenghentikan proses secepatnya...")
            current_log("Menutup seluruh browser dan membatalkan antrean...")
        except Exception:
            pass

    cleanup_all_drivers()

def perform_manual_login(login_url, log):
    if stop_event.is_set():
        return None
    log("Cepetan loginnya, jangan kelamaan, cepetan!!!")
    options = ChromeOptions()
    options.add_argument("--start-maximized")
    
    # Evasion to prevent bot detection
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    service = ChromeService()
    if sys.platform == "win32":
        service.creation_flags = 0x08000000  # CREATE_NO_WINDOW
        
    driver = webdriver.Chrome(service=service, options=options)
    with drivers_lock:
        drivers.append(driver)
    
    # Bypass navigator.webdriver detection
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })
    
    try:
        if stop_event.is_set():
            return None
        driver.get(login_url)
        if stop_event.is_set():
            return None
        # Tunggu sampai elemen table muncul (tanda bahwa login sudah berhasil dan masuk ke halaman)
        wait = WebDriverWait(driver, 120)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
        if stop_event.is_set():
            return None
        time.sleep(1)
        if stop_event.is_set():
            return None
        
        # Ambil cookies untuk diberikan ke background browser
        cookies = driver.get_cookies()
        return cookies
    finally:
        with drivers_lock:
            if driver in drivers:
                drivers.remove(driver)
        try:
            driver.quit()
        except:
            pass

def get_driver():
    # Jika thread sudah memiliki driver, periksa apakah driver tersebut masih valid (terdaftar di global drivers).
    # Jika driver tidak terdaftar (karena telah ditutup oleh stop_download), bersihkan referensi lokalnya.
    if hasattr(thread_local, "driver"):
        driver_valid = False
        with drivers_lock:
            if thread_local.driver in drivers:
                driver_valid = True
        if not driver_valid:
            if hasattr(thread_local, "driver"):
                delattr(thread_local, "driver")

    if not hasattr(thread_local, "driver"):
        if stop_event.is_set():
            raise RuntimeError("Download dihentikan (stop_event aktif).")
        options = ChromeOptions()
        options.add_argument("--headless=new") # Jalan di background
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--remote-allow-origins=*")
        options.add_argument("--window-size=1920,1080")
        
        # Evasion to prevent bot detection
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        
        service = ChromeService()
        if sys.platform == "win32":
            service.creation_flags = 0x08000000  # CREATE_NO_WINDOW
            
        driver = webdriver.Chrome(service=service, options=options)
        with drivers_lock:
            drivers.append(driver)
            
        try:
            # Bypass navigator.webdriver detection
            driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
            })
            
            # Pindah ke domain root dulu agar selenium mengizinkan injeksi cookies
            driver.get("https://dev1.sikap.lampungprov.go.id/favicon.ico")
            
            global shared_cookies
            for cookie in shared_cookies:
                driver.add_cookie(cookie)
                
            thread_local.driver = driver
        except Exception as e:
            # Jika gagal inisialisasi dan bukan karena stop_event, bersihkan driver ini
            if not stop_event.is_set():
                with drivers_lock:
                    if driver in drivers:
                        drivers.remove(driver)
                try:
                    driver.quit()
                except:
                    pass
            raise e
    return thread_local.driver

def process_pegawai(obj, base_url, dir_rekap, log, progress_info=None):
    if stop_event.is_set():
        return False

    id_peg = obj["id"]
    nama = sanitize_filename(obj["name"])
    bidang = sanitize_filename(obj.get("bidang", "Lainnya"))

    folder = os.path.join(dir_rekap, bidang)
    os.makedirs(folder, exist_ok=True)

    out = os.path.join(folder, f"{nama}.pdf")

    if stop_event.is_set():
        return False

    log(f"Download {nama}...")

    try:
        if stop_event.is_set():
            return False
        driver = get_driver()
        wait = WebDriverWait(driver, 30)

        # Sebelum driver.get()
        if stop_event.is_set():
            return False
        driver.get(base_url.format(id=id_peg))
        # Sesudah driver.get()
        if stop_event.is_set():
            return False

        # Sebelum wait.until()
        if stop_event.is_set():
            return False
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
        # Sesudah wait.until()
        if stop_event.is_set():
            return False

        time.sleep(1)
        if stop_event.is_set():
            return False

        # Sebelum execute_cdp_cmd()
        if stop_event.is_set():
            return False
        pdf = driver.execute_cdp_cmd("Page.printToPDF", {"scale": 0.5})
        # Sesudah execute_cdp_cmd()
        if stop_event.is_set():
            return False

        # Sebelum write PDF
        if stop_event.is_set():
            return False
        with open(out, "wb") as f:
            f.write(base64.b64decode(pdf["data"]))
        # Sesudah write PDF
        if stop_event.is_set():
            return False

        if progress_info:
            with progress_info["lock"]:
                progress_info["completed"] += 1
                completed = progress_info["completed"]
                total = progress_info["total"]
                remaining = total - completed
            log(f"{completed}/{total}: Download {nama} selesai.")
        else:
            log(f"Download {nama} selesai.")

        return True
    except Exception as e:
        # Bersihkan driver lokal yang crash/detached agar tidak dipakai ulang
        if hasattr(thread_local, "driver"):
            dead_driver = thread_local.driver
            delattr(thread_local, "driver")
            with drivers_lock:
                if dead_driver in drivers:
                    drivers.remove(dead_driver)
            try:
                dead_driver.quit()
            except Exception:
                pass

        if stop_event.is_set():
            # Jika dihentikan oleh user, abaikan log error Selenium agar tidak membingungkan
            return False
        if progress_info:
            with progress_info["lock"]:
                progress_info["completed"] += 1
                completed = progress_info["completed"]
                total = progress_info["total"]
                remaining = total - completed
            log(f"Error download {nama}: {str(e)}. (Sudah didownload: {completed}/{total}, Belum: {remaining})")
        else:
            log(f"Error download {nama}: {str(e)}")
        return False

def download_rekap(excel_pegawai, dir_rekap, bulan, tahun, log, max_workers=5):
    global current_log, drivers, shared_cookies
    current_log = log
    stop_event.clear()
    
    with drivers_lock:
        drivers.clear()

    wb = load_workbook(excel_pegawai, data_only=True)
    ws = wb.active

    # Ambil header pada baris pertama
    headers = [cell.value for cell in ws[1]]

    # Ubah setiap baris menjadi dictionary
    items = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue

        items.append(dict(zip(headers, row)))

    if not items:
        log("Tidak ada data pegawai untuk didownload.")
        return

    BASE_URL = f"https://dev1.sikap.lampungprov.go.id/app/cetak-laporan/data-harian-bulanan-pegawai?bulan={bulan}&tahun={tahun}&id_peg={{id}}"
    
    # Login menggunakan pegawai pertama
    first_url = BASE_URL.format(id=items[0]["id"])
    try:
        shared_cookies = perform_manual_login(first_url, log)
        if stop_event.is_set() or shared_cookies is None:
            log("Download dihentikan oleh pengguna.")
            return
        log("Login terdeteksi!")
    except Exception as e:
        if stop_event.is_set():
            log("Download dihentikan oleh pengguna.")
            return
        log(f"Gagal memverifikasi login atau waktu tunggu habis. Pesan error: {e}")
        return

    if stop_event.is_set():
        log("Download dihentikan oleh pengguna.")
        return

    # max_workers sudah diterima dari parameter (default 5)
    total_files = len(items)
    log(f"Total file yang akan didownload: {total_files} file.")
    log(f"Memulai download. {max_workers}x lebih cepat dari biasanya!")

    progress_info = {
        "lock": threading.Lock(),
        "completed": 0,
        "total": total_files
    }

    start_time = time.time()

    try:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = []
            for obj in items:
                if stop_event.is_set():
                    break
                futures.append(
                    executor.submit(
                        process_pegawai,
                        obj,
                        BASE_URL,
                        dir_rekap,
                        log,
                        progress_info
                    )
                )

            # Menunggu seluruh task selesai, periksa stop_event secara berkala
            while futures and not stop_event.is_set():
                if all(f.done() for f in futures):
                    break
                time.sleep(0.5)

            if stop_event.is_set():
                # Batalkan task yang belum berjalan
                for f in futures:
                    f.cancel()
                
                # Shutdown executor secara non-blocking dengan cancel_futures (Python 3.9+)
                if sys.version_info >= (3, 9):
                    executor.shutdown(wait=False, cancel_futures=True)
                else:
                    executor.shutdown(wait=False)
            else:
                # Ambil hasil untuk melempar exception jika ada worker yang crash saat mode normal
                for f in futures:
                    try:
                        f.result()
                    except Exception:
                        pass

    finally:
        # Tutup seluruh browser secara aman dan hilangkan sisa proses Chrome
        cleanup_all_drivers()

    if stop_event.is_set():
        log("Download dihentikan oleh pengguna.")
    else:
        duration = time.time() - start_time
        log(f"Download selesai dalam {duration:.2f} detik.")