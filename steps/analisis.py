import os
import re
import shutil
from datetime import datetime
import pdfplumber

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

from steps.work_calendar import WorkCalendar, parse_tanggal_indonesia


# =========================================================
# FUNGSI GLOBAL (WAJIB DI LUAR)
# =========================================================
def clean_name(val):
    if not val:
        return ""
    s = re.sub(r"[.,]", "", str(val)).upper()
    return re.sub(r"\s+", " ", s).strip()


def klasifikasi_keterlambatan(jam, tanggal, kalender):

    if not jam or not tanggal:
        return None

    jam_norm = kalender.get_jam_masuk(tanggal)

    jam_dt = datetime.combine(tanggal, jam.time())
    jam_norm_dt = datetime.combine(tanggal, jam_norm)

    selisih_menit = (jam_dt - jam_norm_dt).total_seconds() / 60

    if selisih_menit <= 0:
        return None
    elif selisih_menit <= 15:
        return "<15"
    elif selisih_menit < 31:
        return "<30"
    elif selisih_menit <= 60:
        return "<60"
    else:
        return ">60"


# =========================================================
# ANALISIS KEHADIRAN
# =========================================================
def analisis_kehadiran(dir_rekap, template_excel, output_excel, log, json_kalender_kerja=""):

    kalender = WorkCalendar(json_kalender_kerja)
    if json_kalender_kerja:
        if kalender.rules:
            log(f"✅ Berhasil memuat kalender kerja kustom: {os.path.basename(json_kalender_kerja)}")
            for r in kalender.rules:
                log(f"   • {r['nama']}: {r['mulai']} s/d {r['selesai']} (Jam Masuk: {r['jam_masuk']})")
        else:
            log(f"⚠ File kalender kerja ditentukan ({os.path.basename(json_kalender_kerja)}) tetapi kosong atau gagal dimuat. Menggunakan default ASN 07:30.")

    def parse_jam(teks):
        if not teks:
            return None

        # PDF kadang pakai titik
        teks = re.sub(r"[.,]", ":", teks.strip())

        try:
            return datetime.strptime(teks, "%H:%M:%S")
        except:
            return None

    def is_zero_or_missing(value):
        if not value:
            return True
        value = value.replace(",", ".").strip()
        try:
            return float(value) == 0.0
        except:
            return True
    
    def durasi_adalah_nol(isi):

        # contoh pada PDF: 6.5   0,00   0.00
        durasi_match = re.findall(r"\d{1,2}[.,]\d{1,2}", isi)

        if not durasi_match:
            return True  # tidak ada durasi → anggap tidak hadir

        # biasanya angka terakhir adalah durasi kerja
        durasi_str = durasi_match[-1].replace(",", ".")

        try:
            return float(durasi_str) == 0.0
        except:
            return True
        
    # === BACA PDF ===
    hasil_rekap = {}

    # Kumpulkan semua file PDF untuk menghitung total & melaporkan progres
    pdf_files = []
    for root, dirs, files in os.walk(dir_rekap):
        for file in files:
            if file.lower().endswith(".pdf"):
                pdf_files.append((root, file))

    total_files = len(pdf_files)
    log(f"Menemukan {total_files} file PDF rekap kehadiran.")

    for idx, (root, file) in enumerate(pdf_files, 1):
        from steps.download import stop_event
        if stop_event.is_set():
            log("Analisis kehadiran dibatalkan oleh pengguna.")
            raise SystemExit()

        nama = os.path.splitext(file)[0].strip()
        file_path = os.path.join(root, file)

        # Log progres setiap 10 file agar user tahu aplikasi sedang bekerja
        if idx % 10 == 0 or idx == 1 or idx == total_files:
            log(f"  - [{idx}/{total_files}] Membaca & menganalisis: {nama}")

        with pdfplumber.open(file_path) as pdf:
            text = "\n".join([p.extract_text() or "" for p in pdf.pages])

            # normalisasi
            text = re.sub(r"\n+", "\n", text).strip()

            # =====================================================
            # DETEKSI TANGGAL INDONESIA
            # =====================================================
            tanggal_regex = r"(\d{1,2}\s+(?:Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember)\s+20\d{2})"

            # Normalisasi spasi di tanggal agar "10 April 2026" dan "10 April\n2026"
            # dianggap sama, supaya Total HK tidak bertambah hanya karena format teks PDF.
            tanggal_list_raw = re.findall(tanggal_regex, text)
            tanggal_list = [re.sub(r"\s+", " ", t.strip()) for t in tanggal_list_raw]
            tanggal_unik = list(dict.fromkeys(tanggal_list))

            bagian = re.split(tanggal_regex, text)

            blok_tanggal = {}
            for i in range(1, len(bagian), 2):
                tanggal_raw = bagian[i]
                tanggal = re.sub(r"\s+", " ", tanggal_raw.strip())
                isi = bagian[i+1] if i+1 < len(bagian) else ""

                # Jika tanggal yang sama muncul lagi (mis. di keterangan WFH),
                # gabungkan isi-nya supaya jam & durasi tidak hilang dan tidak terbaca TMK.
                if tanggal in blok_tanggal:
                    blok_tanggal[tanggal] += " " + isi
                else:
                    blok_tanggal[tanggal] = isi

            keterlambatan = {"<15": 0, "<30": 0, "<60": 0, ">60": 0}
            tmk = 0

            # =====================================================
            # PROSES PER TANGGAL
            # =====================================================
            for tgl, isi in blok_tanggal.items():
                isi = re.split(r"Total\s*\(Dalam", isi, flags=re.IGNORECASE)[0]
                # konversi "18 Februari 2026" -> date
                tanggal_obj = parse_tanggal_indonesia(tgl)
                if tanggal_obj is None:
                    continue

                # abaikan dinas luar, cuti dll
                isi_low = isi.lower()
                if any(x in isi_low for x in ["dinas luar", "tubel", "tubbel", "cuti", "banding"]):
                    continue

                # =================================================
                # CARI JAM MASUK (AMBIL WAKTU PERTAMA SAJA)
                # =================================================
                jam_masuk = None
                jm = re.search(r"(\d{2}:\d{2}:\d{2})", isi)

                if jm:
                    jam_masuk = parse_jam(jm.group(1))

                # CARI DURASI
                dur = re.search(r"(\d{1,2}[.,]\d{1,2})\s+(\d{1,2}[.,]\d{1,2})", isi)
                durasi = None
                if dur:
                    durasi = dur.group(2).replace(",", ".")
                
                if not jam_masuk or is_zero_or_missing(durasi):
                    tmk += 1
                    continue

                if durasi_adalah_nol(isi):
                    tmk += 1
                    continue

                # =================================================
                # HITUNG KETERLAMBATAN BERDASARKAN KALENDER
                # =================================================
                kategori = klasifikasi_keterlambatan(jam_masuk, tanggal_obj, kalender)

                if kategori:
                    keterlambatan[kategori] += 1

            data_dict = {
                "<15": keterlambatan["<15"],
                "<30": keterlambatan["<30"],
                "<60": keterlambatan["<60"],
                ">60": keterlambatan[">60"],
                "TMK": tmk,
                "HariKerja": len(tanggal_unik)
            }
            hasil_rekap[nama.strip()] = data_dict
            c_name = clean_name(nama)
            if c_name:
                hasil_rekap[c_name] = data_dict

    # =====================================================
    # TULIS EXCEL
    # =====================================================
    log("Menulis hasil analisis ke Excel...")
    shutil.copy(template_excel, output_excel)

    wb = load_workbook(template_excel, keep_links=True)

    # Cari sheet "DISIPLIN" secara case-insensitive
    ws = None
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().upper() == "DISIPLIN":
            ws = wb[sheet_name]
            break

    # Fallback: jika hanya ada 1 sheet, gunakan sheet itu (apapun namanya)
    if ws is None:
        if len(wb.sheetnames) == 1:
            ws = wb[wb.sheetnames[0]]
            log(f"  ⚠ Sheet '{wb.sheetnames[0]}' ditemukan, otomatis digunakan sebagai sheet DISIPLIN.")
        else:
            sheet_list = ", ".join(wb.sheetnames) or "(tidak ada sheet)"
            raise ValueError(
                f"Sheet 'DISIPLIN' tidak ditemukan di file template Excel.\n"
                f"Sheet yang tersedia: {sheet_list}\n"
                f"Silakan rename salah satu sheet menjadi 'DISIPLIN', atau gunakan template Excel yang benar."
            )

    # Auto-rename sheet ke DISIPLIN jika namanya berbeda
    if ws.title.strip().upper() != "DISIPLIN":
        ws.title = "DISIPLIN"

    kolom = {}
    for col in range(1, ws.max_column+1):
        val = ws.cell(1, col).value
        if val:
            kolom[val.strip()] = col

    map_col = {
        "Nama": kolom.get("NAMA"),
        "TMK": kolom.get("TMK"),
        "<15": kolom.get("< 15 menit"),
        "<30": kolom.get("< 30 menit"),
        "<60": kolom.get("< 60 menit"),
        ">60": kolom.get("> 60 menit"),
        "HariKerja": kolom.get("Total HK")
    }

    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )

    for row in range(2, ws.max_row+1):
        nama = ws.cell(row=row, column=map_col["Nama"]).value
        if not nama:
            continue

        nama_raw = str(nama).strip()
        nama_clean = clean_name(nama_raw)

        data = None
        if nama_raw in hasil_rekap:
            data = hasil_rekap[nama_raw]
        elif nama_clean in hasil_rekap:
            data = hasil_rekap[nama_clean]

        if data:

            # Warnai seluruh sel di baris ini dengan hijau muda sebagai penanda absensi terproses
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = green_fill

            for k in ["<15", "<30", "<60", ">60", "TMK"]:
                cell = ws.cell(row, map_col[k])
                cell.value = data[k]

                if data[k] > 0:
                    cell.fill = red_fill

                cell.border = border

            ws.cell(row, map_col["HariKerja"]).value = data["HariKerja"]

    log(f"Menulis dan menyimpan hasil ke Excel: {output_excel}")
    wb.save(output_excel)
    wb.close()

    log("Analisis selesai")