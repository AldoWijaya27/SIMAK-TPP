import pandas as pd
from openpyxl import load_workbook
import shutil


def clean_nip(val):
    if val is None or pd.isna(val):
        return ""
    if isinstance(val, float):
        val = f"{val:.0f}"
    import re
    s = str(val).strip()
    return re.sub(r"\D", "", s)


def clean_name(val):
    if not val or pd.isna(val):
        return ""
    import re
    s = re.sub(r"[.,]", "", str(val)).upper()
    return re.sub(r"\s+", " ", s).strip()


def merge_ekin_apel(file_sumber, file_utama, output_path):
    try:
        shutil.copy(file_utama, output_path)
        df_sumber = pd.read_excel(file_sumber)
        df_sumber.columns = df_sumber.columns.str.strip()

        required_cols = ["NIP", "Predikat Kinerja", "TMA", 'TMA Lain']
        for col in required_cols:
            if col not in df_sumber.columns:
                raise ValueError(f"Kolom '{col}' tidak ditemukan di file sumber.")

        # Buat mapping berdasarkan NIP murni & Nama murni (sebagai fallback)
        mapping_nip = {}
        mapping_nama = {}

        col_nama_sumber = "NAMA" if "NAMA" in df_sumber.columns else ("Nama" if "Nama" in df_sumber.columns else None)

        for _, row in df_sumber.iterrows():
            nip_clean = clean_nip(row["NIP"])
            data_dict = {
                "Predikat Kinerja": row["Predikat Kinerja"],
                "TMA": row["TMA"],
                "TMA Lain": row["TMA Lain"]
            }
            if nip_clean and len(nip_clean) >= 5:
                mapping_nip[nip_clean] = data_dict

            if col_nama_sumber:
                nama_clean = clean_name(row[col_nama_sumber])
                if nama_clean:
                    mapping_nama[nama_clean] = data_dict

        # ==========================================================
        # 3️⃣ Load file output pakai openpyxl (bukan pandas!)
        # ==========================================================
        wb = load_workbook(output_path)
        ws = wb["DISIPLIN"]  # jika sheet tertentu, bisa diganti wb["NamaSheet"]

        # ==========================================================
        # 4️⃣ Identifikasi posisi kolom berdasarkan header
        # ==========================================================
        header_row = 1
        headers = {}

        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=header_row, column=col).value
            if header_value:
                headers[str(header_value).strip()] = col

        if "NIP" not in headers:
            raise ValueError("Kolom 'NIP' tidak ditemukan di file utama.")

        if "Predikat Kinerja" not in headers:
            raise ValueError("Kolom 'Predikat Kinerja' tidak ditemukan di file utama.")

        if "TMA" not in headers:
            raise ValueError("Kolom 'TMA' tidak ditemukan di file utama.")
        
        if "TMA Lain" not in headers:
            raise ValueError("Kolom 'TMA Lain' tidak ditemukan di file utama.")

        col_nip = headers["NIP"]
        col_nama = headers.get("NAMA") or headers.get("Nama")
        col_predikat = headers["Predikat Kinerja"]
        col_tma = headers["TMA"]
        col_tma_lain = headers["TMA Lain"]

        # ==========================================================
        # 5️⃣ Update nilai berdasarkan NIP / Nama (tanpa merusak rumus lain)
        # ==========================================================
        updated_count = 0

        for row in range(2, ws.max_row + 1):
            nip_cell = ws.cell(row=row, column=col_nip).value
            nip_clean = clean_nip(nip_cell)

            target_data = None
            if nip_clean in mapping_nip:
                target_data = mapping_nip[nip_clean]
            elif col_nama:
                nama_cell = ws.cell(row=row, column=col_nama).value
                nama_clean = clean_name(nama_cell)
                if nama_clean in mapping_nama:
                    target_data = mapping_nama[nama_clean]

            if target_data:
                ws.cell(row=row, column=col_predikat).value = target_data["Predikat Kinerja"]
                ws.cell(row=row, column=col_tma).value = target_data["TMA"]
                ws.cell(row=row, column=col_tma_lain).value = target_data["TMA Lain"]
                updated_count += 1

        # Save file setelah update data NIP
        wb.save(output_path)
        wb.close()

        # ==========================================================
        # 6️⃣ Rekalkulasi seluruh rumus resmi menggunakan Microsoft Excel background
        # ==========================================================
        from steps.excel_csv import recalculate_excel_via_win32
        recalculate_excel_via_win32(output_path)

        # Update juga file_utama (Disiplin_TPP.xlsx) agar pengguna melihat perubahan di kedua file
        try:
            shutil.copy(output_path, file_utama)
        except Exception:
            pass

        return True, f"Berhasil update {updated_count} data berdasarkan NIP."

    except Exception as e:
        return False, str(e)
