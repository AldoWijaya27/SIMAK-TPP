import os
import re
import calendar
from datetime import datetime, date, time

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from steps.work_calendar import parse_tanggal_indonesia


# =========================================================
# KONSTANTA
# =========================================================
JAM_BATAS_APEL = time(7, 45, 0)  # Batas jam masuk apel: 07.45.00

# Warna formatting sel
FILL_TK     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Merah muda
FILL_WFH    = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Biru muda
FILL_LIBUR  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Abu-abu
FILL_HEADER = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")  # Hijau tua (header)
FILL_KETERANGAN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Hijau muda (keterangan CT/DL/dll)

FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_TK     = Font(name="Calibri", size=10, bold=True, color="9C0006")
FONT_WFH    = Font(name="Calibri", size=9, italic=True, color="1F4E79")
FONT_LIBUR  = Font(name="Calibri", size=9, italic=True, color="595959")
FONT_KETERANGAN = Font(name="Calibri", size=9, italic=True, color="375623")
FONT_TITLE  = Font(name="Calibri", size=12, bold=True)

# Nama bulan Indonesia
NAMA_BULAN = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember"
]

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
)

# Nama hari Indonesia (0=Senin, 6=Minggu)
NAMA_HARI = ["Sen", "Sel", "Rab", "Kam", "Jum", "Sab", "Min"]


# =========================================================
# FUNGSI: Baca Data Pegawai dari Excel (Data Pegawai)
# =========================================================
def _baca_data_pegawai(excel_pegawai, log):
    """
    Membaca file Excel Data Pegawai dan mengembalikan list of dict
    berisi: name, nip, jabatan, bidang.
    Juga mengembalikan daftar bidang unik sesuai urutan pertama kali muncul.
    """
    from openpyxl import load_workbook

    wb = load_workbook(excel_pegawai, data_only=True)
    ws = wb.active

    # Ambil header baris 1
    headers_raw = [cell.value for cell in ws[1]]
    headers = []
    for h in headers_raw:
        if h:
            headers.append(str(h).strip())
        else:
            headers.append("")

    # Cari indeks kolom yang dibutuhkan (case-insensitive)
    headers_lower = [h.lower() for h in headers]

    def find_col(keywords):
        for kw in keywords:
            for i, h in enumerate(headers_lower):
                if kw in h:
                    return i
        return None

    col_nama   = find_col(["name", "nama"])
    col_nip    = find_col(["nip"])
    col_jabatan = find_col(["jabatan"])
    col_bidang = find_col(["bidang"])

    if col_nama is None:
        raise ValueError("Kolom 'name/NAMA' tidak ditemukan di file Data Pegawai.")
    if col_bidang is None:
        raise ValueError("Kolom 'bidang/BIDANG' tidak ditemukan di file Data Pegawai.")

    pegawai_list = []
    bidang_order = []
    seen_names = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue

        nama = str(row[col_nama]).strip() if row[col_nama] else ""
        if not nama:
            continue

        nama_lower = nama.lower()
        if nama_lower in seen_names:
            log(f"  ⚠ Peringatan: Nama ganda dilewati -> {nama}")
            continue
        seen_names.add(nama_lower)

        nip = str(row[col_nip]).strip() if col_nip is not None and row[col_nip] else ""
        jabatan = str(row[col_jabatan]).strip() if col_jabatan is not None and row[col_jabatan] else ""
        bidang = str(row[col_bidang]).strip() if row[col_bidang] else "Lainnya"

        pegawai_list.append({
            "nama": nama,
            "nip": nip,
            "jabatan": jabatan,
            "bidang": bidang,
        })

        if bidang not in bidang_order:
            bidang_order.append(bidang)

    wb.close()
    log(f"  Membaca {len(pegawai_list)} pegawai dari {len(bidang_order)} bidang.")
    return pegawai_list, bidang_order


# =========================================================
# FUNGSI: Ekstrak Jam Masuk dari PDF per Tanggal
# =========================================================
def _ekstrak_jam_masuk_dari_pdf(file_path):
    """
    Membaca file PDF rekap kehadiran dan mengembalikan dict:
      { date_obj: datetime_jam_masuk_or_None, ... }
    Hanya tanggal yang ditemukan di PDF yang dimasukkan.
    """
    hasil = {}

    with pdfplumber.open(file_path) as pdf:
        text = "\n".join([p.extract_text() or "" for p in pdf.pages])

    # normalisasi
    text = re.sub(r"\n+", "\n", text).strip()

    # Deteksi tanggal Indonesia
    tanggal_regex = r"(\d{1,2}\s+(?:Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember)\s+20\d{2})"

    bagian = re.split(tanggal_regex, text)

    blok_tanggal = {}
    for i in range(1, len(bagian), 2):
        tanggal_raw = bagian[i]
        tanggal = re.sub(r"\s+", " ", tanggal_raw.strip())
        isi = bagian[i+1] if i+1 < len(bagian) else ""

        if tanggal in blok_tanggal:
            blok_tanggal[tanggal] += " " + isi
        else:
            blok_tanggal[tanggal] = isi

    for tgl, isi in blok_tanggal.items():
        # Potong di "Total (Dalam..."
        isi = re.split(r"Total\s*\(Dalam", isi, flags=re.IGNORECASE)[0]

        tanggal_obj = parse_tanggal_indonesia(tgl)
        if tanggal_obj is None:
            continue

        # Ekstrak keterangan khusus (Hanya CT dan DL yang tampil teks, keterangan lain seperti izin/sakit/tubel dikosongkan selnya)
        isi_low = isi.lower()
        if "cuti" in isi_low:
            hasil[tanggal_obj] = "CT"
            continue
        elif any(x in isi_low for x in ["dinas luar", "perjalanan dinas", "perjadin"]):
            hasil[tanggal_obj] = "DL"
            continue
        elif any(x in isi_low for x in ["tubel", "tubbel", "tugas belajar", "sakit", "izin", "ijin", "banding"]):
            hasil[tanggal_obj] = "__EXCUSED__"
            continue

        # Cari jam masuk (waktu pertama saja)
        jam_masuk = None
        jm = re.search(r"(\d{2}:\d{2}:\d{2})", isi)
        if jm:
            teks_jam = re.sub(r"[.,]", ":", jm.group(1).strip())
            try:
                jam_masuk = datetime.strptime(teks_jam, "%H:%M:%S").time()
            except Exception:
                jam_masuk = None

        hasil[tanggal_obj] = jam_masuk

    return hasil


# =========================================================
# FUNGSI: Tentukan Status Apel untuk 1 Tanggal
# =========================================================
def _status_apel(tanggal_obj, jam_masuk_time, is_public_holiday=False):
    """
    Menentukan status apel berdasarkan hari dan jam masuk.

    Returns:
        tuple: (status_text, is_counted_tk)
        - status_text: "" (hadir/kosong), "TK", "CT", "DL"
        - is_counted_tk: True jika dihitung ke Total TK
    """
    weekday = tanggal_obj.weekday()  # 0=Senin, 4=Jumat, 5=Sabtu, 6=Minggu

    if weekday == 4:  # Jumat — WFH, sel kosong tapi tetap diwarnai
        return "__WFH__", False
    elif weekday in (5, 6):  # Sabtu, Minggu — LIBUR, sel kosong tapi tetap diwarnai
        return "__LIBUR__", False
    elif is_public_holiday:  # Libur nasional — sel kosong tapi tetap diwarnai
        return "__LIBUR__", False
    else:
        # Senin-Kamis: jadwal apel
        if jam_masuk_time == "__EXCUSED__":
            # Izin, Sakit, Tubel, Banding, dll: sel kosong, tidak dihitung TK
            return "", False
        elif isinstance(jam_masuk_time, str):
            # CT dan DL -> tampilkan teks, tidak dihitung TK
            return jam_masuk_time, False
        elif jam_masuk_time is None:
            # Jam masuk kosong / tanpa keterangan -> sel kosong, tidak dihitung TK (dianggap masuk/hadir)
            return "", False
        elif jam_masuk_time <= JAM_BATAS_APEL:
            # Hadir tepat waktu
            return "", False
        else:
            # Terlambat apel (> 07.45)
            return "TK", True


# =========================================================
# FUNGSI UTAMA: Rekap Kehadiran Apel
# =========================================================
def _tulis_sheet_rekap_apel(ws, bidang, daftar_peg, tanggal_bulan, pdf_map, tanggal_ada_di_pdf, nama_bulan_text, tahun_int):
    """
    Mengisi satu worksheet dengan tabel rekap kehadiran apel untuk suatu bidang.
    Mengembalikan dict mapping nip -> total_tk.
    """
    # ── Header Judul (Row 1-4) ──
    ws.cell(row=1, column=1, value="REKAPITULASI APEL").font = FONT_TITLE
    ws.cell(row=2, column=1, value=f"BIDANG/KPH : {bidang}").font = FONT_TITLE
    ws.cell(row=3, column=1, value=f"BULAN : {nama_bulan_text} {tahun_int}").font = FONT_TITLE

    # ── Tabel Header (Row 5) ──
    header_row = 5
    kolom_tetap = ["No", "NIP", "Nama", "Jabatan"]
    kolom_tanggal = [str(tgl.day) for tgl in tanggal_bulan]
    headers = kolom_tetap + kolom_tanggal + ["Total TK"]

    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # ── Data Rows (mulai dari row 6) ──
    col_start_tanggal = len(kolom_tetap) + 1
    col_total_tk = len(headers)
    data_start_row = header_row + 1
    nip_to_tk = {}

    for row_offset, peg in enumerate(daftar_peg):
        row_idx = data_start_row + row_offset
        nama = peg["nama"]
        nip = peg["nip"]
        jabatan = peg["jabatan"]

        ws.cell(row=row_idx, column=1, value=row_offset + 1).font = FONT_NORMAL
        ws.cell(row=row_idx, column=1).alignment = ALIGN_CENTER
        ws.cell(row=row_idx, column=1).border = THIN_BORDER

        ws.cell(row=row_idx, column=2, value=nip).font = FONT_NORMAL
        ws.cell(row=row_idx, column=2).alignment = ALIGN_LEFT
        ws.cell(row=row_idx, column=2).border = THIN_BORDER

        ws.cell(row=row_idx, column=3, value=nama).font = FONT_NORMAL
        ws.cell(row=row_idx, column=3).alignment = ALIGN_LEFT
        ws.cell(row=row_idx, column=3).border = THIN_BORDER

        ws.cell(row=row_idx, column=4, value=jabatan).font = FONT_NORMAL
        ws.cell(row=row_idx, column=4).alignment = ALIGN_LEFT
        ws.cell(row=row_idx, column=4).border = THIN_BORDER

        from utils import sanitize_filename
        jam_data = (
            pdf_map.get(nama)
            or pdf_map.get(nama.lower())
            or pdf_map.get(sanitize_filename(nama).lower(), {})
        )

        total_tk = 0

        for tgl_idx, tgl in enumerate(tanggal_bulan):
            col = col_start_tanggal + tgl_idx
            jam_masuk = jam_data.get(tgl, None)

            is_public_holiday = (tgl not in tanggal_ada_di_pdf)
            status, is_tk = _status_apel(tgl, jam_masuk, is_public_holiday)

            if status == "__WFH__":
                cell = ws.cell(row=row_idx, column=col, value="")
                cell.fill = FILL_WFH
                cell.font = FONT_WFH
            elif status == "__LIBUR__":
                cell = ws.cell(row=row_idx, column=col, value="")
                cell.fill = FILL_LIBUR
                cell.font = FONT_LIBUR
            elif status == "TK":
                cell = ws.cell(row=row_idx, column=col, value="x")
                cell.fill = FILL_TK
                cell.font = FONT_TK
                total_tk += 1
            elif status:
                cell = ws.cell(row=row_idx, column=col, value=status)
                cell.fill = FILL_KETERANGAN
                cell.font = FONT_KETERANGAN
            else:
                cell = ws.cell(row=row_idx, column=col, value="")
                cell.font = FONT_NORMAL

            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER

        cell_total = ws.cell(row=row_idx, column=col_total_tk, value=total_tk)
        cell_total.font = Font(name="Calibri", size=10, bold=True)
        cell_total.alignment = ALIGN_CENTER
        cell_total.border = THIN_BORDER
        if total_tk > 0:
            cell_total.fill = FILL_TK

        if nip:
            nip_clean = str(nip).replace(" ", "").strip()
            nip_to_tk[nip_clean] = total_tk

    # ── Atur lebar kolom ──
    ws.column_dimensions[get_column_letter(1)].width = 5    # No
    ws.column_dimensions[get_column_letter(2)].width = 22   # NIP
    ws.column_dimensions[get_column_letter(3)].width = 30   # Nama
    ws.column_dimensions[get_column_letter(4)].width = 25   # Jabatan

    for tgl_idx in range(len(tanggal_bulan)):
        col = col_start_tanggal + tgl_idx
        ws.column_dimensions[get_column_letter(col)].width = 5

    ws.column_dimensions[get_column_letter(col_total_tk)].width = 10
    ws.freeze_panes = f"E{data_start_row}"

    return nip_to_tk


# =========================================================
# FUNGSI UTAMA: Rekap Kehadiran Apel
# =========================================================
def rekap_kehadiran_apel(dir_rekap, excel_pegawai, bulan, tahun, output_path, log, ekin_apel_excel=None):
    """
    Membuat file Excel Rekap Kehadiran Apel per Bidang dan File Master.
    Juga mengisikan Total TK ke kolom TMA di file Ekin & Apel.

    Args:
        dir_rekap        (str): Folder REKAP KEHADIRAN DITANDATANGANI
        excel_pegawai    (str): Path ke file Excel Data Pegawai
        bulan            (str): Bulan periode (mis. "06")
        tahun            (int): Tahun periode (mis. 2026)
        output_path      (str): Path output (opsional/legacy)
        log              (callable): Fungsi log ke GUI
        ekin_apel_excel  (str): Path ke file Excel Data Ekin & Apel (opsional)
    """
    log("── Memulai proses Rekap Kehadiran Apel ──")

    bulan_int = int(bulan)
    tahun_int = int(tahun)

    # 1. Baca Data Pegawai
    log("  [1/4] Membaca data pegawai...")
    pegawai_list, bidang_order = _baca_data_pegawai(excel_pegawai, log)

    if not pegawai_list:
        log("❌ Tidak ada data pegawai ditemukan.")
        return

    # 2. Susun daftar tanggal dalam bulan
    jumlah_hari = calendar.monthrange(tahun_int, bulan_int)[1]
    tanggal_bulan = [date(tahun_int, bulan_int, d) for d in range(1, jumlah_hari + 1)]
    log(f"  Periode: Bulan {bulan_int}, Tahun {tahun_int} ({jumlah_hari} hari)")

    # 3. Baca semua PDF Rekap Kehadiran
    log("  [2/4] Membaca file PDF rekap kehadiran...")
    pdf_map = {}
    pdf_files = []
    for root, dirs, files in os.walk(dir_rekap):
        for file in files:
            if file.lower().endswith(".pdf"):
                pdf_files.append(os.path.join(root, file))

    total_pdf = len(pdf_files)
    log(f"  Menemukan {total_pdf} file PDF.")

    for idx, file_path in enumerate(pdf_files, 1):
        from steps.download import stop_event
        if stop_event.is_set():
            log("Proses Rekap Apel dibatalkan oleh pengguna.")
            raise SystemExit()

        nama_file = os.path.splitext(os.path.basename(file_path))[0].strip()

        if idx % 10 == 0 or idx == 1 or idx == total_pdf:
            log(f"    [{idx}/{total_pdf}] Membaca: {nama_file}")

        try:
            jam_data = _ekstrak_jam_masuk_dari_pdf(file_path)
            pdf_map[nama_file] = jam_data
            pdf_map[nama_file.lower()] = jam_data
            from utils import sanitize_filename
            pdf_map[sanitize_filename(nama_file).lower()] = jam_data
        except Exception as e:
            log(f"  ⚠ Gagal membaca PDF {nama_file}: {e}")

    tanggal_ada_di_pdf = set()
    for jam_data in pdf_map.values():
        tanggal_ada_di_pdf.update(jam_data.keys())

    # 4. Buat File Per-Bidang & Workbook Master
    log("  [3/4] Membuat file Excel rekap apel...")

    pegawai_per_bidang = {}
    for peg in pegawai_list:
        bidang = peg["bidang"]
        if bidang not in pegawai_per_bidang:
            pegawai_per_bidang[bidang] = []
        pegawai_per_bidang[bidang].append(peg)

    nama_bulan_text = NAMA_BULAN[bulan_int] if 1 <= bulan_int <= 12 else str(bulan_int)
    output_files = []
    all_nip_to_tk = {}

    for bidang in bidang_order:
        from steps.download import stop_event
        if stop_event.is_set():
            log("Proses Rekap Apel dibatalkan oleh pengguna.")
            raise SystemExit()

        daftar_peg = pegawai_per_bidang.get(bidang, [])
        if not daftar_peg:
            continue

        log(f"    Membuat rekap: {bidang} ({len(daftar_peg)} pegawai)")

        # --- File Per-Bidang ---
        wb_bidang = Workbook()
        ws_bidang = wb_bidang.active
        ws_bidang.title = "Rekap Apel"
        nip_tk_dict = _tulis_sheet_rekap_apel(
            ws_bidang, bidang, daftar_peg, tanggal_bulan,
            pdf_map, tanggal_ada_di_pdf, nama_bulan_text, tahun_int
        )
        all_nip_to_tk.update(nip_tk_dict)

        folder_bidang = os.path.join(dir_rekap, bidang)
        os.makedirs(folder_bidang, exist_ok=True)
        from utils import sanitize_filename
        nama_file_output = sanitize_filename(f"Rekap_Apel_{bidang}_{bulan}_{tahun}.xlsx")
        path_output_bidang = os.path.join(folder_bidang, nama_file_output)
        wb_bidang.save(path_output_bidang)
        wb_bidang.close()
        output_files.append(path_output_bidang)

    # 5. Auto-fill TMA ke File Ekin & Apel
    path_ekin_output = os.path.join(dir_rekap, sanitize_filename(f"Ekin_Apel_Terisi_{bulan}_{tahun}.xlsx"))
    if ekin_apel_excel and os.path.exists(ekin_apel_excel):
        log("  [4/4] Mengisikan Total TK ke kolom TMA pada file Ekin & Apel...")
        try:
            wb_ekin = load_workbook(ekin_apel_excel)
            ws_ekin = wb_ekin["DISIPLIN"] if "DISIPLIN" in wb_ekin.sheetnames else wb_ekin.active

            # Cari posisi kolom NIP dan TMA
            headers = {}
            for col in range(1, ws_ekin.max_column + 1):
                hv = ws_ekin.cell(row=1, column=col).value
                if hv:
                    headers[str(hv).strip().upper()] = col

            col_nip = headers.get("NIP")
            col_tma = headers.get("TMA")

            if col_nip and col_tma:
                updated_count = 0
                for r in range(2, ws_ekin.max_row + 1):
                    nip_val = ws_ekin.cell(row=r, column=col_nip).value
                    if nip_val:
                        nip_clean = str(nip_val).replace(" ", "").strip()
                        if nip_clean in all_nip_to_tk:
                            ws_ekin.cell(row=r, column=col_tma).value = all_nip_to_tk[nip_clean]
                            updated_count += 1

                wb_ekin.save(path_ekin_output)
                wb_ekin.close()
                log(f"  ✅ Berhasil update {updated_count} pegawai di file Ekin & Apel → {path_ekin_output}")
            else:
                log("  ⚠ Kolom NIP atau TMA tidak ditemukan pada file Ekin & Apel.")
                wb_ekin.close()
        except Exception as e:
            log(f"  ⚠ Gagal mengisi file Ekin & Apel: {e}")
    else:
        log("  [4/4] Membuat file Ekin_Apel_Terisi otomatis sesuai format template...")
        try:
            wb_new = Workbook()
            ws_new = wb_new.active
            ws_new.title = "DISIPLIN"

            # Headers persis template Gambar 2
            headers_template = ["No.", "Nama Pegawai", "NIP", "Predikat Kinerja", "TMA", "TMA Lain"]
            ws_new.append(headers_template)

            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # Style Header Row 1
            for col_idx in range(1, 7):
                cell = ws_new.cell(row=1, column=col_idx)
                cell.font = Font(name="Calibri", size=11, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            seen_nips = set()
            no_counter = 1
            for p in pegawai_list:
                nip_raw = str(p.get("nip", "")).strip()
                nip_clean = nip_raw.replace(" ", "").strip()
                if not nip_clean or nip_clean.lower() in ["nan", "none", "-", "0"]:
                    continue
                if nip_clean in seen_nips:
                    continue
                seen_nips.add(nip_clean)

                nama_val = p.get("nama") or p.get("name") or ""
                tk_count = all_nip_to_tk.get(nip_clean, 0)

                row_data = [
                    no_counter,
                    nama_val,
                    nip_raw,
                    "Baik/Sangat Baik",
                    tk_count,
                    0
                ]
                ws_new.append(row_data)

                row_idx = ws_new.max_row
                for col_idx in range(1, 7):
                    c = ws_new.cell(row=row_idx, column=col_idx)
                    c.font = Font(name="Calibri", size=11)
                    c.border = thin_border
                    if col_idx in [1, 5, 6]:
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        c.alignment = Alignment(horizontal="left", vertical="center")

                no_counter += 1

            for col in ws_new.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws_new.column_dimensions[col_letter].width = max(max_len + 4, 12)

            wb_new.save(path_ekin_output)
            wb_new.close()
            log(f"  ✅ File Ekin & Apel Terisi otomatis dibuat ({no_counter - 1} pegawai) → {path_ekin_output}")
        except Exception as e:
            log(f"  ⚠ Gagal membuat file Ekin & Apel Terisi otomatis: {e}")

    log(f"✅ Rekap Kehadiran Apel selesai! {len(output_files)} file bidang disimpan.")
