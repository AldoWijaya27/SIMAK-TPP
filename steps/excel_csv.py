import os
import pandas as pd

def recalculate_excel_via_win32(file_path):
    """
    Meminta Microsoft Excel (via COM API Windows) untuk menghitung dan menyimpan 
    seluruh rumus asli di dalam file .xlsx secara otomatis di background.
    """
    import sys, os
    if sys.platform != "win32" or not os.path.exists(file_path):
        return False
    try:
        import pythoncom
        import win32com.client as win32
        
        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        abs_path = os.path.abspath(file_path)
        wb = excel.Workbooks.Open(abs_path)
        wb.Save()
        wb.Close()
        excel.Quit()
        pythoncom.CoUninitialize()
        return True
    except Exception as e:
        print(f"Peringatan: Recalculate via Win32 Excel gagal/skip: {e}")
        return False


def recalculate_disiplin_sheet(ws):
    """
    Fungsi ini dulunya menghitung ulang rumus secara hardcode di Python.
    Sekarang di-bypass agar 100% menggunakan hasil rumus murni dari Microsoft Excel.
    """
    pass



def format_rupiah(val):
    if val is None or pd.isna(val) or str(val).strip() == "":
        return "0"
    try:
        s = str(val).strip()
        import re
        s = re.sub(r"(?i)rp\.?\s*", "", s)
        if "." in s and "," in s:
            s = s.replace(".", "").replace(",", ".")
        elif "." in s:
            parts = s.split(".")
            if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3 and parts[0].isdigit()):
                s = s.replace(".", "")
        elif "," in s:
            s = s.replace(",", ".")

        n = int(round(float(s)))
        return f"{n:,}".replace(",", ".")
    except (ValueError, TypeError):
        return str(val)


def format_persen(val):
    if val is None or pd.isna(val) or str(val).strip() == "":
        return "0,00"
    try:
        s = str(val).strip().replace("%", "")
        import re
        s = re.sub(r"(?i)rp\.?\s*", "", s)
        f = float(s)
        formatted = f"{f:.2f}".replace(".", ",")
        return formatted
    except (ValueError, TypeError):
        return str(val)


def release_file_lock_if_needed(file_path):
    """
    Jika file sedang terkunci di Windows oleh WINWORD.EXE atau wps.exe (sisa Mail Merge sebelumnya),
    secara otomatis mematikan proses background Word/WPS agar file dapat ditulis ulang.
    """
    import sys, os
    if sys.platform != "win32" or not os.path.exists(file_path):
        return
    import subprocess, time
    try:
        with open(file_path, "a"):
            pass
        return
    except IOError:
        pass

    try:
        subprocess.run(
            ["taskkill", "/F", "/IM", "WINWORD.EXE"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            creationflags=0x08000000
        )
        subprocess.run(
            ["taskkill", "/F", "/IM", "kwps.exe"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            creationflags=0x08000000
        )
        subprocess.run(
            ["taskkill", "/F", "/IM", "wps.exe"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            creationflags=0x08000000
        )
        time.sleep(0.5)
    except Exception:
        pass


def excel_sheet_disiplin_ke_csv(file_excel, output_folder):
    file_excel = os.path.abspath(file_excel)
    output_folder = os.path.abspath(output_folder)

    filename = os.path.splitext(os.path.basename(file_excel))[0] + ".csv"
    output_csv = os.path.join(output_folder, filename)

    # Lepaskan penguncian file jika ada proses background Word yang masih memegang file
    release_file_lock_if_needed(file_excel)
    release_file_lock_if_needed(output_csv)

    # Hitung dan simpan seluruh rumus resmi menggunakan Microsoft Excel background
    recalculate_excel_via_win32(file_excel)

    from openpyxl import load_workbook
    wb = load_workbook(file_excel, data_only=True)
    sheet_name = "DISIPLIN" if "DISIPLIN" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]


    # Evaluasi kolom TTE untuk CSV
    tte_col = None
    nip_p_col = None
    jab_p_col = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "").strip().lower()
        if h == "tte":
            tte_col = c
        elif h == "nip pimpinan":
            nip_p_col = c
        elif h == "jabatan pimpinan":
            jab_p_col = c

    if tte_col:
        auth_nips = {"196902051989101002", "198204182006042012"}
        for r in range(2, ws.max_row + 1):
            nip_p = str(ws.cell(r, nip_p_col).value or "").replace(" ", "").strip() if nip_p_col else ""
            jab_p = str(ws.cell(r, jab_p_col).value or "").strip().upper() if jab_p_col else ""

            is_auth = (nip_p in auth_nips) or (("KEPALA DINAS" in jab_p or "SEKRETARIS" in jab_p) and "SEKRETARIS DAERAH" not in jab_p)
            if is_auth:
                ws.cell(r, tte_col).value = "${ttd_pengirim}"
            else:
                ws.cell(r, tte_col).value = ""

    # Catat indeks baris yang berwarna hijau (absensi ter-download/terproses)
    valid_data_indices = set()
    for r in range(2, ws.max_row + 1):
        is_green = False
        for c in range(1, min(ws.max_column + 1, 10)):
            fill = ws.cell(r, c).fill
            if fill and fill.fill_type == "solid" and fill.start_color:
                color_str = str(getattr(fill.start_color, "rgb", "") or "").upper()
                if "E2EFDA" in color_str:
                    is_green = True
                    break
        if is_green:
            valid_data_indices.add(r - 2)  # index 0-based untuk DataFrame (karena data_rows[1:])

    # Ambil nilai numerik untuk DataFrame CSV
    data_rows = list(ws.values)
    headers_list = data_rows[0] if data_rows else []
    df = pd.DataFrame(data_rows[1:], columns=headers_list) if len(data_rows) > 1 else pd.DataFrame()

    wb.close()

    # Filter DataFrame CSV: Hanya sertakan pegawai yang barisnya berwarna hijau (absensi terproses)
    if not df.empty and valid_data_indices:
        valid_indices_in_df = [idx for idx in range(len(df)) if idx in valid_data_indices]
        if valid_indices_in_df:
            df = df.iloc[valid_indices_in_df].copy()



    # Format kolom nominal rupiah dengan pemisah titik ribuan (misal 4.949.000)
    currency_cols = [
        "TPP Asli", "jumlah TP", "Tambahan 20%", "TPP Kotor",
        "PPh21", "BPJS", "jumlah bersih", "zakat", "TPP bersih",
        "pengurangan disiplin", "jumlah potongan"
    ]
    for col in currency_cols:
        if col in df.columns:
            df[col] = df[col].apply(format_rupiah)

    # Format persentase dengan 2 desimal koma (misal 60,00%, 40,00%, 2,00%, 98,00%)
    pct_cols = [
        "Skor Kinerja (%)", "Skor Kehadiran (%)", "TMK Persen",
        "Persentase Total", "Skor Total (%)", "Persentase Kinerja",
        "TMA Persen", "TMA Lain Persen", "Persen a", "Persen b",
        "Persen c", "Persen d", "persentase hadir", "Skor Tidak Disiplin"
    ]
    for col in pct_cols:
        if col in df.columns:
            df[col] = df[col].apply(format_persen)

    if "tte" in df.columns:
        df["tte"] = df["tte"].fillna("")

    try:
        df.to_csv(output_csv, sep=';', encoding='utf-8', index=False)
    except PermissionError:
        release_file_lock_if_needed(output_csv)
        try:
            df.to_csv(output_csv, sep=';', encoding='utf-8', index=False)
        except PermissionError:
            raise Exception(f"File CSV '{os.path.basename(output_csv)}' sedang dibuka di Microsoft Excel atau aplikasi lain. Silakan tutup jendela Excel/Word Anda terlebih dahulu lalu coba lagi.")
